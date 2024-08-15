# coding=utf-8
from openpyxl.utils import get_column_letter
import xlwings as xw
from xlwings import App
import pandas as pd
import datetime
from openpyxl.utils import column_index_from_string
import tkinter as tk
from tkinter import ttk  #下拉框
import tkinter.messagebox #弹出框
from tkinter import *
#from xlutils.copy import copy
import os
import shutil
from tkinter import filedialog
current_file_path = os.getcwd()
os.chdir(current_file_path)
from sys import path
path.append(current_file_path)
# from ExcelOpt import *
import xlsxwriter
import numpy as np
import warnings
from PIL import ImageTk
import time
import traceback
warnings.filterwarnings('ignore')
# 创建窗口
window = tk.Tk()
window.title('项目型物控表工具')  # 窗口的标题
window.geometry('600x800')  # 窗口的大小
window.iconbitmap('软件附带文件\头像.ico')
frame=tk.Canvas(window,width=620,height=600,background='silver',scrollregion=(0,0,1500,1000))
roll=Scrollbar(window,orient='vertical',command=frame.yview)
#frame.pack(fill="both",side='right')
frame['yscrollcommand']=roll.set
roll.pack(side=RIGHT, fill=Y)
frame.pack(side=TOP, fill=Y, expand=True)
image_file = ImageTk.PhotoImage(file=r'软件附带文件\2023.jpg')

image =frame.create_image(300, 0, anchor='n', image=image_file)
image =frame.create_image(300, 210, anchor='n', image=image_file)

image_file1 = ImageTk.PhotoImage(file=r'软件附带文件\海目星.jpg')
image_new =frame.create_image(500, 10, anchor='n', image=image_file1)
#image2 =frame.create_image(300, 1005, anchor='n', image=image_file)

# 1分类
lab_choice = tk.Label(frame,
             text='选择执行方式',  # 标签的文字
             bg='darkred',  # 标签背景颜色
             font=('华文行楷', 15),  # 字体和字体大小
             width=18, height=1,  # 标签长宽(以字符长度计算)
             padx=2,pady=4,anchor='s',fg='white')

frame.create_window(90,25,window=lab_choice)
cv= tkinter.StringVar()  #跟踪变量变化
choice=['手动方式','执行深圳共享盘','执行江门/江苏共享盘']
box_choice=ttk.Combobox(frame,values=choice,font=('华文行楷',15),width=22,textvariable=cv)  #下拉框
#设置默认值
box_choice.current(0)
frame.create_window(320,25,window=box_choice)
window.update()

def mean():
    box_mean=tk.messagebox.askokcancel(title='操作说明',message='选择合适的执行方式（手动版本数据源文件夹内所有文件都需准备好，自动版本需准备好cbmq200，三个版本都需准备好需求项目列表），然后点击执行')
button_mean= tk.Button(frame, text='操作说明',command=mean,width=8, height=1, fg='darkred',bd=4,font=('华文行楷', 12))
frame.create_window(530,110,window=button_mean)
button_execute= tk.Button(frame, text='执行', width=6, height=1, fg='darkred',bd=6,font=('华文行楷', 20))
frame.create_window(300,110,window=button_execute)

def execute():
    start_time = time.time()
    screm = tk.Text(frame, bg='white',  # 标签背景颜色
                    font=('微软雅黑', 12),  # 字体和字体大小
                    width=60, height=28,  # 标签长宽(以字符长度计算)
                    )
    frame.create_window(300, 450, window=screm)

    screm.insert(INSERT,'请确认“数据源”文件夹内：apmr929、cpmq500、cinr002、cbmq200、需求项目表,是否准备完毕', '\n')
    screm.insert(INSERT, '\n本次运行时间大约1-5分钟,请勿关闭本窗口！', '\n')
    screm.insert(INSERT, '\n已完成{:.0%}'.format(0), '\n')
    window.update()
    try:
#手动
        if box_choice.get()=='手动方式':
            screm.insert(INSERT, '\n读取500...', '\n')
            window.update()
            filePath1 = '数据源/cpmq500'
            file_name1 = os.listdir(filePath1)
            for i in range(len(file_name1)):
                if str(file_name1[i]).count('~$') == 0 and str(file_name1[i]).count('LEBG') == 0:
                    with open(os.path.join(filePath1,  str(file_name1[i])), encoding='gb18030', errors='ignore') as f:
                        # 再解决部分报错行如 ParserError：Error tokenizing data.C error:Expected 2 fields in line 407,saw 3.
                        report_500 = pd.read_csv(f, sep=',', error_bad_lines=False, low_memory=False, thousands=',')
                    #report_500 = pd.read_csv(filePath1 + '/' + str(file_name1[i]), encoding='gb18030', sep=',',low_memory=False)
                if str(file_name1[i]).count('LEBG') > 0:
                    with open(os.path.join(filePath1, str(file_name1[i])), encoding='utf-8', errors='ignore') as f:
                        report_500 = pd.read_csv(f, sep=',',low_memory=False)

            filePath2 = '数据源/apmr929'
            screm.insert(INSERT, '\n读取929...', '\n')
            window.update()
            file_name2 = os.listdir(filePath2)
            if os.listdir(filePath2):
                for i in range(len(file_name2)):
                    if str(file_name2[i]).count('~$') == 0 and str(file_name2[i]).count('LEBG') == 0:
                        report_929 = pd.read_excel(filePath2 + '/' + str(file_name2[i]), header=3)
                        report_929 = report_929.iloc[:-2]
                    if str(file_name1[i]).count('LEBG') > 0:
                        report_929 = pd.read_csv(filePath2 + '/' + str(file_name2[i]), encoding='utf-8', sep=',',low_memory=False)
            else:
                report_929= pd.DataFrame(columns=["项目编号",'最新需求时间','采购员名称','核价采购员名称','备注说明','请购单号','项次','交货地址栏','项目阶段'])

            filePath3 = '数据源/cbmq200'
            screm.insert(INSERT, '\n读取200...', '\n')
            window.update()
            file_name3 = os.listdir(filePath3)
            for i in range(len(file_name3)):
                if str(file_name3[i]).count('~$') == 0:
                    report_200 = pd.read_excel(filePath3 + '/' + str(file_name3[i]))
            filePath4 = '数据源/cinr002'
            screm.insert(INSERT, '\n读取002...', '\n')
            window.update()
            file_name4 = os.listdir(filePath4)
            if os.listdir(filePath2):
                for i in range(len(file_name4)):
                    if str(file_name4[i]).count('~$') == 0 and str(file_name4[i]).count('LEBG') == 0:
                        report_002 = pd.read_excel(filePath4 + '/' + str(file_name4[i]), header=3)
                        report_002 = report_002.iloc[:-2]
                    if str(file_name4[i]).count('LEBG') > 0:
                        report_002 = pd.read_csv(filePath4 + '/' + str(file_name4[i]), encoding='utf-8', sep=',',low_memory=False)
            else:
                report_002= pd.DataFrame(columns=['单据编号','录入日期','申请人员','人员名称','资料创建日','数据审核日','变更类型','变更类型说明','料件编号','品名','规格','库位','库位说明','变更前-库存管理特征','变更前-库存单位','变更前单位名称','变更后-库存管理特征','变更后-库存单位','变更后单位名称','变更数量','备注'])
    #江门江苏
        if box_choice.get() == '执行江门/江苏共享盘':
            it = pd.read_excel('数据源/需求项目列表.xlsx')
            it_1 = str(it['项目号'][0])[:3]
            time_now = time.strftime("%Y%m%d", time.localtime())
            time_now_h = time.strftime("%H", time.localtime())
            if it_1.count('JM') == 1:
                # filePath1 = '数据源/cpmq500'
                if int(time_now_h) < 13:
                    file_time = time_now + 'am'
                    fil1 = r'\\192.168.32.10\lebg计划管理科\部门公共\ERPdata\topprd\LEBGJM\\' + file_time  #\\192.168.32.10\lebg计划管理科\部门公共\ERPdata\topprd\LEBGJM
                    if os.path.exists(fil1):
                        report_500 = pd.read_csv(fil1 + r'\\cpmq500-LEBGJM.csv', encoding='utf-8', sep=',',low_memory=False)
                        report_929 = pd.read_csv(fil1 + r'\\apmr929-LEBGJM.csv', encoding='utf-8', sep=',',low_memory=False)
                        report_002 = pd.read_csv(fil1 + r'\\cinr002-LEBGJM.csv', encoding='utf-8', sep=',',low_memory=False)
                    else:
                        print('共享盘数据最新数据疑似未传输，请核实')
                        screm.insert(INSERT,'\n共享盘数据最新数据疑似未传输，请核实', '\n')
                if int(time_now_h) >= 13:
                    file_time = time_now + 'pm'
                    fil1 = r'\\192.168.32.10\lebg计划管理科\部门公共\ERPdata\topprd\LEBGJM\\' + file_time
                    if os.path.exists(fil1):
                        report_500 = pd.read_csv(fil1 + r'\\cpmq500-LEBGJM.csv', encoding='utf-8', sep=',',
                                                 low_memory=False)
                        report_929 = pd.read_csv(fil1 + r'\\apmr929-LEBGJM.csv', encoding='utf-8', sep=',',
                                                 low_memory=False)
                        report_002 = pd.read_csv(fil1 + r'\\cinr002-LEBGJM.csv', encoding='utf-8', sep=',',
                                                 low_memory=False)
                    else:
                        print('共享盘数据最新数据疑似未传输，请核实')
                        screm.insert(INSERT, '\n共享盘数据最新数据疑似未传输，请核实', '\n')
            if it_1.count('JS') == 1:
                # filePath1 = '数据源/cpmq500'
                if int(time_now_h) < 13:
                    file_time = time_now + 'am'
                    fil1 = r'\\10.4.1.225\lebg - 计划管理\ERPdata\topprd\LEBGJS\\' + file_time
                    if os.path.exists(fil1):
                        report_500 = pd.read_csv(fil1 + r'\\cpmq500-LEBGJS.csv', encoding='utf-8', sep=',',low_memory=False)
                        report_929 = pd.read_csv(fil1 + r'\\apmr929-LEBGJS.csv', encoding='utf-8', sep=',',low_memory=False)
                        report_002 = pd.read_csv(fil1 + r'\\cinr002-LEBGJS.csv', encoding='utf-8', sep=',',low_memory=False)
                    else:
                        print('共享盘数据最新数据疑似未传输，请核实')
                        screm.insert(INSERT, '\n共享盘数据最新数据疑似未传输，请核实', '\n')
                if int(time_now_h) >= 13:
                    file_time = time_now + 'pm'
                    fil1 = r'\\10.4.1.225\lebg - 计划管理\ERPdata\topprd\LEBGJS\\' + file_time
                    if os.path.exists(fil1):
                        report_500 = pd.read_csv(fil1 + r'\\cpmq500-LEBGJS.csv', encoding='utf-8', sep=',',low_memory=False)
                        report_929 = pd.read_csv(fil1 + r'\\apmr929-LEBGJS.csv', encoding='utf-8', sep=',',low_memory=False)
                        report_002 = pd.read_csv(fil1 + r'\\cinr002-LEBGJS.csv', encoding='utf-8', sep=',',low_memory=False)
                    else:
                        print('共享盘数据最新数据疑似未传输，请核实')
                        screm.insert(INSERT, '\n共享盘数据最新数据疑似未传输，请核实', '\n')
            filePath3 = '数据源/cbmq200'
            file_name3 = os.listdir(filePath3)
            for i in range(len(file_name3)):
                if str(file_name3[i]).count('~$') == 0:
                    report_200 = pd.read_excel(filePath3 + '/' + str(file_name3[i]))

    #深圳
        if box_choice.get()=='执行深圳共享盘':
            it = pd.read_excel('数据源/需求项目列表.xlsx')
            it_1 = str(it['项目号'][0])[:3]
            time_now = time.strftime("%Y%m%d", time.localtime())
            time_now_h = time.strftime("%H", time.localtime())
            if it_1.count('JM') == 1:
                # filePath1 = '数据源/cpmq500'
                if int(time_now_h) < 13:
                    file_time = time_now + 'am'
                    fil1 = r'\\10.10.80.252\激光及能源行业中心\(05)部门文件夹\(26)PMO\10.ERPdata\topprd\LEBGJM\\' + file_time
                    if os.path.exists(fil1):
                        report_500 = pd.read_csv(fil1 + r'\\cpmq500-LEBGJM.csv', encoding='utf-8', sep=',',
                                                 low_memory=False)
                        report_929 = pd.read_csv(fil1 + r'\\apmr929-LEBGJM.csv', encoding='utf-8', sep=',',
                                                 low_memory=False)
                        report_002 = pd.read_csv(fil1 + r'\\cinr002-LEBGJM.csv', encoding='utf-8', sep=',',
                                                 low_memory=False)
                    else:
                        print('共享盘数据最新数据疑似未传输，请核实')
                        screm.insert(INSERT, '\n共享盘数据最新数据疑似未传输，请核实', '\n')
                if int(time_now_h) >= 13:
                    file_time = time_now + 'pm'
                    fil1 = r'\\10.10.80.252\激光及能源行业中心\(05)部门文件夹\(26)PMO\10.ERPdata\\topprd\LEBGJM\\' + file_time
                    if os.path.exists(fil1):
                        report_500 = pd.read_csv(fil1 + r'\\cpmq500-LEBGJM.csv', encoding='utf-8', sep=',',
                                                 low_memory=False)
                        report_929 = pd.read_csv(fil1 + r'\\apmr929-LEBGJM.csv', encoding='utf-8', sep=',',
                                                 low_memory=False)
                        report_002 = pd.read_csv(fil1 + r'\\cinr002-LEBGJM.csv', encoding='utf-8', sep=',',
                                                 low_memory=False)
                    else:
                        print('共享盘数据最新数据疑似未传输，请核实')
                        screm.insert(INSERT, '\n共享盘数据最新数据疑似未传输，请核实', '\n')
            if it_1.count('JS') == 1:
                # filePath1 = '数据源/cpmq500'
                if int(time_now_h) < 13:
                    file_time = time_now + 'am'
                    fil1 = r'\\10.10.80.252\激光及能源行业中心\(05)部门文件夹\(26)PMO\10.ERPdata\\topprd\LEBGJS\\' + file_time
                    if os.path.exists(fil1):
                        report_500 = pd.read_csv(fil1 + r'\\cpmq500-LEBGJS.csv', encoding='utf-8', sep=',',
                                                 low_memory=False)
                        report_929 = pd.read_csv(fil1 + r'\\apmr929-LEBGJS.csv', encoding='utf-8', sep=',',
                                                 low_memory=False)
                        report_002 = pd.read_csv(fil1 + r'\\cinr002-LEBGJS.csv', encoding='utf-8', sep=',',
                                                 low_memory=False)
                    else:
                        print('共享盘数据最新数据疑似未传输，请核实')
                        screm.insert(INSERT, '\n共享盘数据最新数据疑似未传输，请核实', '\n')
                if int(time_now_h) >= 13:
                    file_time = time_now + 'pm'
                    fil1 = r'\\10.10.80.252\激光及能源行业中心\(05)部门文件夹\(26)PMO\10.ERPdata\\topprd\LEBGJS\\' + file_time
                    if os.path.exists(fil1):
                        report_500 = pd.read_csv(fil1 + r'\\cpmq500-LEBGJS.csv', encoding='utf-8', sep=',',
                                                 low_memory=False)
                        report_929 = pd.read_csv(fil1 + r'\\apmr929-LEBGJS.csv', encoding='utf-8', sep=',',
                                                 low_memory=False)
                        report_002 = pd.read_csv(fil1 + r'\\cinr002-LEBGJS.csv', encoding='utf-8', sep=',',
                                                 low_memory=False)
                    else:
                        print('共享盘数据最新数据疑似未传输，请核实')
                        screm.insert(INSERT, '\n共享盘数据最新数据疑似未传输，请核实', '\n')

            filePath3 = '数据源/cbmq200'
            file_name3 = os.listdir(filePath3)
            for i in range(len(file_name3)):
                if str(file_name3[i]).count('~$') == 0:
                    report_200 = pd.read_excel(filePath3 + '/' + str(file_name3[i]))
        # 这段代码主要执行数据的处理和整理，包括从不同数据源中读取数据、清理和格式化数据、合并数据以及填充缺失值。
        window.update()
        time1 = time.time()
        print('读取报表耗时:%d秒' % (time1 - start_time))
        screm.insert(INSERT,'\n读取报表耗时:%d秒' % (time1 - start_time))
        print('已完成{:.0%}'.format(0.26))
        screm.insert(INSERT,'\n已完成{:.0%}'.format(0.26), '\n')
        window.update()
        # 筛选出需要的列,两张表中各需要32列
        # 筛选出需要的列,两张表中各需要32列
        need=['项目编号', '库存管理特征', '模组名称', '作业编号', '请购日期', '来源单号',
           '物料请购负责人', '状态码', '行状态', '项次', '来源项次','零件类型',
          '料件编号', '品名', '规格', '需求数量',  '品牌','表面处理','采购员姓名','开单人姓名',
          '核价采购员姓名', '采购日期','数据审核日', '采购单号', '供应商名称','未转采购量','是否CEG确认','CEG备注',
              '项目名称', '采购确认日期',
          '供应商回复交期', '已收货量', '已入库量', '仓退数量', '验退数量',
          '未交量', '请购单备注', '采购延期说明','延期进度说明','设计备注','交货地址栏(请购单)','交货地址栏(采购单)','需求日期','最新需求日期','到货日期','物流信息','CEG确认货期(天)','大项目名称','项目阶段','采购确认日期履历', '最新需求日期变更履历','MC负责人','退货快递单号','已检验数量','入库日期','理由码说明','备注说明']

        # 修改名称使两边表格一致
        # 929报表需要新增列
        add_col = ['行状态','采购日期','采购单号','供应商名称','采购确认日期','供应商回复交期','数据审核日','已收货量','已入库量','仓退数量','验退数量','未交量']
        full_col = report_929.columns.tolist() + add_col
        report929 = report_929.reindex(columns=full_col)

        if '项目编号' in list(report_200.columns):
            report_200 = report_200.rename(columns={'项目编号': '项目号'})
        def add_report(item_report,bookname):
            #global report500,report929,report002,report200
            screm.insert(INSERT, '\n开始生成：'+bookname, '\n')
            window.update()
            time1_1=time.time()
            report500 = report_500[report_500['项目编号'].isin(item_report['项目号'])]
            report500.reset_index(drop=True, inplace=True)
            report929 = report_929[report_929['项目编号'].isin(item_report['项目号'])]
            report929.reset_index(drop=True, inplace=True)
            report500 = report500.rename(columns={'数量':'需求数量','核价采购员名称':'核价采购员姓名','请购单需求日期':'需求日期'
                        ,'延期说明':'采购延期说明','备注':'请购单备注','最新需求时间':'最新需求日期'
                        ,'采购人员姓名':'采购员姓名','仓退量':'仓退数量','验退量':'验退数量'})
            report929 = report929.rename(columns={'最新需求时间':'最新需求日期'
                ,'采购员名称':'采购员姓名','核价采购员名称':'核价采购员姓名'
                ,'请购单号':'来源单号','项次':'来源项次','交货地址栏':'交货地址栏(请购单)'})
            report929 = report929.rename(columns={'延期未下单原因说明':'采购延期说明',
                            '请购单号':'来源单号','备注':'请购单备注','项次':'来源项次'})
            report002 = report_002.rename(columns={'变更后-库存管理特征':'库存管理特征'})
            item_report=item_report.rename(columns={'项目齐料时间':'需求齐料时间'})
            project_list=list(item_report['项目号'])
            report002['项目编号']=''

            #提前获取002对应项目
            report002=report002[sum(report002['库存管理特征'].str.contains(project) for project in project_list)>0].reset_index(drop=True)

            for i in  range(len(report002)):
                if str(report002['库存管理特征'][i]).count('-')==1 and len(report002['库存管理特征'][i].split('-')[report002['库存管理特征'][i].count('-')])<=2:
                    report002['项目编号'][i] = report002['库存管理特征'][i]
                if str(report002['库存管理特征'][i]).count('-')==1 and len(report002['库存管理特征'][i].split('-')[report002['库存管理特征'][i].count('-')])>2:
                    report002['项目编号'][i] = report002['库存管理特征'][i].split('-')[0]
                if str(report002['库存管理特征'][i]).count('-')==2:
                    report002['项目编号'][i] = report002['库存管理特征'][i].split('-')[0]+'-'+report002['库存管理特征'][i].split('-')[1]
                if str(report002['库存管理特征'][i]).count('-') == 0:
                    report002['项目编号'][i] =report002['库存管理特征'][i]
            report002['库存管理特征']=report002['项目编号'].astype(str)+'-无模组号'

                # 删除不需要的列
            if '理由码说明' not in report500.columns:
                report500['理由码说明']=''
            for col in report500.columns:
                if col not in need:
                    del report500[col]
            for col in report929.columns:
                if col not in need:
                    del report929[col]  # 删除不需要的列
            window.update()
            # 合并两张报表
            report1 = pd.concat([report500, report929]).reset_index(drop=True) # 合并2张表格
            # 筛选出所需要的行,并重置索引
            report = report1[report1['项目编号'].isin(item_report['项目号'])]


            report200 = report_200[report_200['项目号'].isin(item_report['项目号'])]
            report200.reset_index(drop=True, inplace=True)
            # 002筛选出所需要的行,并重置索引
            report002 = report002[report002['项目编号'].isin(item_report['项目号'])]
            report002.reset_index(drop=True, inplace=True)

            # 查看有缺失值的列，要求无缺失值的列出现了缺失值时，给出提醒
            exam_col = ['项目编号', '库存管理特征', '请购日期','需求日期']
            for col in exam_col:
                if any(report[col].isnull()):
                    print('请补全929/500表中有缺失值的列:' + col)
                    screm.insert(INSERT, '\n请补全929/500表中有缺失值的列:' + col)
            exam_col = ['项数', '项目号', '需求齐料时间', '设备数量', '计划出货日期',
                        '计划下单总项数', '设计下单完成日期']
            for col in exam_col:
                if any(item_report[col].isnull()):
                    print('请补全需求项目表中有缺失值的列:' + col)
                    screm.insert(INSERT, '\n请补全需求项目表中有缺失值的列:' + col)
            window.update()


            # 合并后的报表需要再额外新增14列
            report = report.reindex(columns=
                                    ['项目编号', '库存管理特征', '作业编号', '模组名称', 'P工序', '系列', '设备','项目阶段',
                                     '负责人', '请购日期', '请购天数', '来源单号','来源项次',
                                      '物料请购负责人', '状态码', '行状态', '料件编号',
                                     '品名', '规格', '需求数量', '未转采购量','品牌', '需求日期', '最新需求日期', '零件类型', '表面处理', '开单人姓名', '核价采购员姓名',
                                     '采购日期','数据审核日', '采购单号','项次', '供应商名称', '采购确认日期', '供应商回复交期',
                                     '已收货量', '已入库量', '仓退数量', '验退数量', '未交量',
                                      '当前状态', '风险等级', '类别分类','交期差异天数', '设计延期天数', '超3天未下单', '交期判断','到货日期','是否大件物料','物流信息','大项目名称',
                                     '采购延期说明','延期进度说明', '是否CEG确认', 'CEG备注', 'CEG确认货期(天)','设计备注','交货地址栏(请购单)','交货地址栏(采购单)', '请购单备注','采购确认日期履历', '最新需求日期变更履历','MC负责人','理由码说明','退货快递单号','已检验数量','入库日期','备注说明'])

            # 区分文本列、日期列和数值列
            str_col = ['项目编号','库存管理特征','零件类型','系列','设备','项目阶段','负责人',
                    'P工序','作业编号','模组名称','表面处理','来源单号',
                    '物料请购负责人','状态码','行状态','料件编号',
                    '品名','规格','品牌','开单人姓名','核价采购员姓名',
                    '采购单号','供应商名称','风险等级','类别分类','当前状态',
                    '超3天未下单','交期判断','采购延期说明','延期进度说明','是否CEG确认','CEG备注','设计备注','交货地址栏(请购单)','交货地址栏(采购单)','请购单备注','物流信息','大项目名称','采购确认日期履历', '最新需求日期变更履历','MC负责人','理由码说明','退货快递单号','备注说明']


            int_col = ['请购天数','项次','来源项次','需求数量','已收货量','已入库量',
                        '仓退数量','验退数量','未交量','交期差异天数','设计延期天数','CEG确认货期(天)','未转采购量','已检验数量']
            date_col = ['请购日期','采购日期','采购确认日期','供应商回复交期','需求日期','最新需求日期','数据审核日','到货日期','入库日期']

            item_report_str_col = ['项目号', '系列', '设备名称', '项目负责人']
            item_report_int_col = ['设备数量', '计划下单总项数', '项数']
            item_report_date_col = ['需求齐料时间', '计划出货日期', '设计下单完成日期']

            # module_report_date_col = ['P工序物料需求时间']
            ###############委外批次
            asfp400_use = asfp400[asfp400['项目编号'].isin(item_report['项目号'])].reset_index(drop=True)
            asfp400_use = asfp400_use.fillna('')

            report = pd.concat([report, asfp400_use]).reset_index(drop=True)
            report = report.reset_index(drop=True)

            # 文本列填充空文本，转为字符串
            report['作业编号'] = report['作业编号'].fillna('无模组号')
            report.loc[report['作业编号']==" ",'作业编号'] ='无模组号'
            for i in  range(len(report)):
                if str(report['库存管理特征'][i]).count('-')>str(report['项目编号'][i]).count('-')and report['作业编号'][i]=='无模组号' and str(report['库存管理特征'][i])[-1]!='-':
                    report['作业编号'][i]=report['库存管理特征'][i].split('-')[report['库存管理特征'][i].count('-')]
                else:
                    next
            report['作业编号'] = report['作业编号'].fillna('无模组号')
            report.loc[report['作业编号']=="无模组号",'模组名称'] ='无作业编号'
            report[str_col] = report[str_col].fillna('').astype(str)
            item_report[item_report_str_col] =item_report[item_report_str_col].fillna('').astype(str)

            # 数值列填充0
            report[int_col] = report[int_col].fillna(0)
            report.loc[report['未转采购量']>0,'需求数量']=report['未转采购量']
            item_report[item_report_int_col] =item_report[item_report_int_col].fillna(0)
            item_report_cop=item_report.copy()
            # 日期列填充1990/01/01，转为pd.Timestamp

            default_date = '1990/01/01'
            report[date_col]=report[date_col].fillna(default_date)
            report['请购日期'] = pd.to_datetime(report["请购日期"],errors='coerce')
            report['采购日期'] = pd.to_datetime(report["采购日期"],errors='coerce')
            report['采购确认日期'] = pd.to_datetime(report["采购确认日期"],errors='coerce')
            report['需求日期'] = pd.to_datetime(report["需求日期"],errors='coerce')
            report['供应商回复交期'] = pd.to_datetime(report["供应商回复交期"],errors='coerce')
            report['数据审核日'] = pd.to_datetime(report["数据审核日"], errors='coerce')
            report['到货日期'] = pd.to_datetime(report["到货日期"], errors='coerce')
            report['最新需求日期'] = pd.to_datetime(report["最新需求日期"],errors='coerce')
            report['入库日期'] = pd.to_datetime(report["入库日期"], errors='coerce')

            #report['CEG确认日期'] = pd.to_datetime(report["CEG确认日期"],errors='coerce')
            report[date_col]=report[date_col].fillna(default_date).astype(str)
            for date_i in date_col:
                for i in  range(len(report)):
                    if int(report[date_i][i][0])==4:
                        delta = pd.Timedelta(str(int(report[date_i][i]))+'days')
                        report[date_i][i] = pd.to_datetime('1899-12-30') + delta
                    if str(report[date_i][i][0:4])=='1970':
                        delta1 = pd.Timedelta(str(int(report[date_i][i][-5:]))+'days')
                        report[date_i][i] = pd.to_datetime('1899-12-30') + delta1
                    else:
                        next
            report['请购日期'] = pd.to_datetime(report["请购日期"],errors='coerce')
            report['采购日期'] = pd.to_datetime(report["采购日期"],errors='coerce')
            report['采购确认日期'] = pd.to_datetime(report["采购确认日期"],errors='coerce')
            report['需求日期'] = pd.to_datetime(report["需求日期"],errors='coerce')
            report['供应商回复交期'] = pd.to_datetime(report["供应商回复交期"],errors='coerce')
            report['数据审核日'] = pd.to_datetime(report["数据审核日"], errors='coerce')
            report['到货日期'] = pd.to_datetime(report["到货日期"], errors='coerce')
            report['最新需求日期'] = pd.to_datetime(report["最新需求日期"],errors='coerce')
            report['入库日期'] = pd.to_datetime(report['入库日期'], errors='coerce')

            report['使用交期']=report['采购确认日期']##############混合采购确认交期与供应商回复交期
            report.loc[report['采购确认日期']==pd.Timestamp(1990, 1, 1),'使用交期']=report['供应商回复交期']

            report = report.rename(columns={'项次': '采购项次'})
            #report['CEG确认日期'] = pd.to_datetime(report["CEG确认日期"],errors='coerce')
            # 日期列填充pd.Timestamp(1990, 1, 1)，转为pd.Timestamp
            default_date = pd.Timestamp(1990, 1, 1)
            item_report[item_report_date_col] = item_report[item_report_date_col].fillna(default_date)
            for col in item_report_date_col:
                item_report[col]= pd.to_datetime(item_report[col], errors='coerce')

            # 日期列填充空文本，转为datetime和文本的混合
            # module_report[module_report_date_col] = \
            #     module_report[module_report_date_col].fillna('')
            item_report_prj=item_report

            time2 = time.time()
            print('%s数据预处理耗时:%d秒' % (bookname,time2-time1_1))
            screm.insert(INSERT, '\n%s数据预处理耗时:%d秒' % (bookname,time2-time1_1))
            print('已完成{:.0%}'.format(0.27))
            screm.insert(INSERT, '\n已完成{:.0%}'.format(0.27))
            window.update()

            # 处理物控表数据列
            #report.loc[:, '序号'] = range(1, len(report) + 1)
            report['库存管理特征'] = report['项目编号'] + '-' + report['作业编号']
            report['请购单备注'] = report['请购单备注'].replace('http', '链接：http', regex=True).astype(str)
            report['设计备注'] = report['设计备注'].replace('http', '链接：http', regex=True).astype(str)


            report.loc[:, '系列'] = vlookup(report.loc[:, '项目编号'],
                                         item_report.loc[:, '项目号'],
                                         item_report.loc[:, '系列'])

            report.loc[:, '设备'] = vlookup(report.loc[:, '项目编号'],
                                         item_report.loc[:, '项目号'],
                                         item_report.loc[:, '设备名称'])

            report.loc[:, '负责人'] = vlookup(report.loc[:, '项目编号'],
                                         item_report.loc[:, '项目号'],
                                         item_report.loc[:, '项目负责人'])

            #类别分类
            report['类别分类']='常规物料'
            report.loc[(report['品名'].str.contains("立柱|电控板|机架|大板|焊接件|横梁|氦检腔体|电控柜|电控箱|直线电机|DD马达|减速机|大理石平台|伺服电机|驱动器|PLC|I/O模块|传感器|易格斯系列线材|太阳线材|工控机|温控器|触摸屏|模组|丝杆|热熔条|拖链|电磁阀|电缸|气缸|蝶阀|阀岛|6轴机器人|滚筒线|干燥炉|超声波焊接机|超声波焊接机焊头焊座|氦检仪|电池内阻测试仪|表面电阻测试仪|电池测试仪器|电池测试仪表|加热板|电柜线材|直线滑轨|笔形气缸|过载断路器|系统电源|方通框架|过辊|收集箱外框|气涨轴|按钮盒|分切刀|激光器|导向轴|过载断路器（施耐德）|齿轮|纠偏器|除尘机|气力输送器|CPU模块|皮带磁力辊|收集箱框架|分切刀小车|振镜|裁断模具|场镜|伺服驱动器|光路|DQ0模组电气物料|直通接头|传感器（松下）|型材外罩")),'类别分类'] = '生产关键物料'
            ##关键物料4.1整改
            bz_uni_list=["穿墙板","磁性过滤器","地脚","电流互感器","电能表","电阻","蝶阀","阀岛","风速仪","浮动接头","干燥炉","高真空气控阀","隔离变压器","工控机","光路","光耦","滚筒线","过载断路器","氦检仪","航空插头","加热板","脚杯","纠偏器","开关电源","漏电断路器","滤波器","平板电脑","热熔条","柔性电缆","三联件","太阳线材","拖链","温控器","显示屏","相机","压力传感器","易格斯","真空泵","真空软管","振镜","CCD","DD马达","FFU","场镜","除尘管","除尘机"]
            report.loc[(report['品名'].isin(bz_uni_list))&(report['零件类型'].str.contains('标准')),'类别分类'] = '生产关键物料'
            ###############################标准键
            ###关键字打头
            report.loc[((report['品名'].str.startswith('光电'))|(report['品名'].str.startswith('激光'))|(report['品名'].str.startswith('耐高温'))|(report['品名'].str.startswith('同步'))|(report['品名'].str.startswith('IO'))|(report['品名'].str.startswith('RSL'))|(report['品名'].str.startswith('安全')))&(report['零件类型'].str.contains('标准')),'类别分类'] = '生产关键物料'
            ###一个关键字收尾
            report.loc[(report['品名'].str[-1:].str.contains('阀|线|针'))&(report['零件类型'].str.contains('标准')), '类别分类'] = '生产关键物料'
            ###两个关键字收尾
            report.loc[(report['品名'].str[-2:].str.contains("板卡|磁铁|弹簧|导轨|电机|堵头|隔环|光源|滚轮|接头|链条|马达|模具|模块|模组|皮带|平键|平台|气缸|切刀|丝杆|铜排|推头|轴承|动轮")) & (report['零件类型'].str.contains('标准')), '类别分类'] = '生产关键物料'
            ###多个关键字收尾
            report.loc[(report['品名'].str[-5:].str.contains("PLC|安装盒|测试仪|传感器|磁力辊|从动轮|电磁阀|电动机|调速阀|发热组件|分支从站|辅助触头组|感应器|固定座|固态继电器|焊接机|缓冲器|机器人|激光器|减速机|接触器|节流阀|控制器|连接器|连接线缆|联轴器|平皮带|驱动器|输送器|温控开关|运动平台|主动轮|专用线束")) & (report['零件类型'].str.contains('标准')), '类别分类'] = '生产关键物料'
            ###############################非标键
            report.loc[(report['品名'].str.contains("安装|大理石|导向|连接|固定|支撑|活动|调整|定位|螺"))&(report['零件类型'].str.contains('标准')==False), '类别分类'] = '生产关键物料'
            ###一个关键字收尾
            report.loc[(report['品名'].str[-1:].str.contains('板|刀|轴')) & (report['零件类型'].str.contains('标准')==False), '类别分类'] = '生产关键物料'
            ###两个关键字收尾
            report.loc[(report['品名'].str[-2:].str.contains("板件|钣金|绑带|保护|材架|侧板|衬套|衬条|撑板|齿轮|大板|带轮|挡板|挡轮|挡片|挡圈|导轨|导轴|导柱|地板|地轨|底板|底座|电缸|垫板|垫块|垫片|垫圈|垫条|顶柱|端块|耳轴|立板|立柱|连杆|链轮|料车|盲板|盘座|配件|器座|腔体|竖板|锁块|推座|托板|外罩|吸板|小车|型轴|压板|压块|压头|载板|支架|支柱|支座|直通|主板|柱销|盖板|隔套|管塞|横板|横梁|滑板|滑轨|滑块|机架|夹板|夹臂|夹块|夹片|夹爪|架板|角座|铰链|脚轮|筋板|卡座|框架|拉板|方块|方通|封板|副轴|转盘|组件|座板")) & (report['零件类型'].str.contains('标准')==False), '类别分类'] = '生产关键物料'
            ###多个关键字收尾
            report.loc[(report['品名'].str[-2:].str.contains("安全柱|保护片|拨料板|不锈钢管|布线板|传动轴|传送件|传送轮|大板件|弹簧轴|导轨座|导向轴|等高螺栓|电机轴|电机座|垫高块|垫高条|调节板|调节块|顶紧块|惰轮轴|法兰座|铝型材架|马达座|门组件|密封垫|密封块|面板塞|平衡杆|气缸转接板|气缸座|气涨轴|器件板|升降板|双链轮|随动块|同步轴|微调块|限位板|限位轴|限位柱|旋转板|旋转座|压力柱|移动块|圆柱头|支持柱|治具板|中转板|轴承座|轴座板|抓料块|转动销|隔热板|过渡板|过渡块|过辊轮|过辊轴|焊接件|焊接体|焊接组件|滑动板|缓冲垫|汇流板|机械手|基准板|基准块|加高块|加强板|加强筋|夹紧臂|夹紧块|减震轴|金属件|拉铆枪|方通架|防尘板|防护板|飞线板|分盘片|转接板|走线板|电控板|电控柜")) & (report['零件类型'].str.contains('标准') == False), '类别分类'] = '生产关键物料'
            #########大件物料
            report['是否大件物料']=''
            report.loc[report['品名'].str.contains("除尘管道|主电控板连接钣金|跨梯|跨梯|跨梯冷水机|喷粉机除尘机|跨梯|冷水机|喷粉机|除尘机|跨梯冷水机|喷粉机|除尘机|跨梯|冷水机喷粉机|除尘机|跨梯|冷水机|喷粉机除尘机|跨梯|搬运机械手|拼接地轨"),'是否大件物料']='是'


            time_delta = list(datetime.datetime.today() - report['请购日期'])
            report.loc[:, '请购天数'] = [item.days if item.days<1000 else 0
                                     for item in time_delta]

            report.loc[:, '当前状态'] = current_statu(report['需求数量'],
                                       report['采购单号'], report['未交量'],
                                       report['仓退数量'], report['验退数量'],
                                       report['已入库量'])
            report.loc[(report['需求数量']>0)&(report['未交量']>0)&(report['使用交期'] != pd.Timestamp(1990, 1, 1))&(datetime.datetime.today().date()>report['使用交期'].dt.date),'当前状态'] = "交付延误"
            report.loc[(report['需求数量']>0)&(report['未交量']>0)&((report['仓退数量']>0)|(report['验退数量']>0)),'当前状态'] = "退回返修"
            report.loc[
                (report['当前状态'] == '退回返修') & ((report['验退数量'] > 0)), '当前状态'] = "来料不良"
            report.loc[
                (report['当前状态'] == '退回返修') & ((report['仓退数量'] > 0)), '当前状态'] = "制程返修"
            report.loc[(report['采购单号'].str.contains('4RA'))&(report['未交量'] > 0) , '当前状态'] = "制程返修"
            report.loc[(report['状态码'].str.contains("留置")),'当前状态'] = '留置'
            report.loc[(report['行状态'].str.contains("留置")),'当前状态'] = '留置'
            report.loc[(report['状态码'].str.contains("结案")),'当前状态'] = '关闭/暂停'
            report.loc[(report['状态码'].str.contains("作废")),'当前状态'] = '关闭/暂停'
            report.loc[(report['行状态'].str.contains("结案")),'当前状态'] = '关闭/暂停'
            report.loc[(report['状态码'].str.contains("短结")),'当前状态'] = '关闭/暂停'
            report.loc[(report['行状态'].str.contains("短结")),'当前状态'] = '关闭/暂停'
            report.loc[(report['需求数量']>0)&(report['未交量']==0)&(report['已入库量']>report['仓退数量']),'当前状态'] = '已入库'
            report.loc[(report['理由码说明'].str.contains('调拨结案采购')) & (report['未交量'] >= 0), '当前状态'] = '已入库'
            #交期差异天数
            report.loc[:, '交期差异天数']=subtract1(report['需求日期'], report['使用交期'],report['最新需求日期'])
            report.loc[(report['需求日期']>=report['使用交期'])&(report['最新需求日期']==pd.Timestamp(1990, 1, 1)),'交期差异天数'] = ''
            report.loc[(report['最新需求日期']>=report['使用交期']),'交期差异天数'] = ''
            report.loc[(report['使用交期']>=report['需求日期'])&(report['需求日期']!=pd.Timestamp(1990, 1, 1))&(datetime.datetime.today().date()>report['使用交期'].dt.date),'交期差异天数'] =(report['需求日期'].dt.date-datetime.datetime.today().date()).apply(lambda x:x / np.timedelta64(1,'D'))
            report.loc[(report['使用交期']>=report['最新需求日期'])&(report['最新需求日期']!=pd.Timestamp(1990, 1, 1))&(datetime.datetime.today().date()>report['使用交期'].dt.date),'交期差异天数'] =(report['需求日期'].dt.date-datetime.datetime.today().date()).apply(lambda x:x / np.timedelta64(1,'D'))
            report.loc[(report['需求日期']>=report['使用交期'])&(report['需求日期']!=pd.Timestamp(1990, 1, 1))&(datetime.datetime.today().date()>report['需求日期'].dt.date),'交期差异天数'] =(report['需求日期'].dt.date-datetime.datetime.today().date()).apply(lambda x:x / np.timedelta64(1,'D'))
            report.loc[(report['最新需求日期']>=report['使用交期'])&(report['最新需求日期']!=pd.Timestamp(1990, 1, 1))&(datetime.datetime.today().date()>report['最新需求日期'].dt.date),'交期差异天数'] =(report['需求日期'].dt.date-datetime.datetime.today().date()).apply(lambda x:x / np.timedelta64(1,'D'))
            report.loc[(report['当前状态']=='已到货')|(report['当前状态']=='关闭/暂停')|(report['当前状态']=='已入库')|(report['当前状态']=='留置')|(report['当前状态']=='未下单'),'交期差异天数'] = ''
            #report['交期差异天数'].astype('str').apply(lambda x:x[:-5]).astype('int32')
            report['交期差异天数']=pd.to_numeric(report['交期差异天数'])

            device_date = vlookup(report.loc[:, '项目编号'],
                                item_report.loc[:, '项目号'],
                                item_report.loc[:, '设计下单完成日期'])
            report['设计延期天数'] = compare_date(report['请购日期'],device_date)
            report['设计延期天数']=pd.to_numeric(report['设计延期天数'])

            report['请购天数']=pd.to_numeric(report['请购天数'])
            report.loc[(report['采购日期']==pd.Timestamp(1990, 1, 1))&(report['请购天数']>3),'超3天未下单'] = '超3天未下单'
            report.loc[(report['当前状态']=='关闭/暂停')|(report['当前状态']=='留置'),'超3天未下单']=''
            #交期判断
            #report['交期判断'] = delivery_judge(report['采购日期'], report['供应商名称'],
                   #            report['当前状态'], report['采购确认日期'],
              #                  report['需求日期'])
            report['需求日期1']=report['需求日期']
            report.loc[(report['最新需求日期']!=pd.Timestamp(1990, 1, 1)),'需求日期1'] =report['最新需求日期']
            report.loc[(report['当前状态'].str.contains('采购中|交付|返修|不良'))&(report['需求日期1']==pd.Timestamp(1990, 1, 1)),'交期判断'] ='无交期'
            report.loc[(report['当前状态'].str.contains('采购|交付|返修|不良'))&(report['需求日期1']!=pd.Timestamp(1990, 1, 1))&(report['使用交期']!=pd.Timestamp(1990, 1, 1)),'交期判断'] ='满足'
            report.loc[(report['当前状态'].str.contains('采购|交付|返修|不良'))&(report['需求日期1']!=pd.Timestamp(1990, 1, 1))&(report['使用交期']==pd.Timestamp(1990, 1, 1)),'交期判断'] ='无交期'
            report.loc[(report['交期差异天数']<0),'交期判断'] = '不满足'
            report.loc[(report['当前状态']=='已到货')|(report['当前状态']=='关闭/暂停')|(report['当前状态']=='已入库')|(report['当前状态']=='留置'),'交期判断'] = '满足'
            report.loc[(report['当前状态']=='未下单'),'交期判断'] = '未下单'
            report.loc[(report['当前状态'].str.contains('采购中|交付|返修|不良'))&(report['需求日期1']==pd.Timestamp(1990, 1, 1)),'交期判断'] ='无交期'
            #风险等级
            report.loc[(report['当前状态'].str.contains('采购中'))&(report['交期判断']=='满足'),'风险等级'] = '低风险'
            report.loc[(report['类别分类']=='生产关键物料')&(report['当前状态'].str.contains('交付|返修|不良'))&(report['交期判断']=='满足'),'风险等级'] = '较大欠料风险'
            report.loc[(report['类别分类']=='常规物料')&(report['当前状态'].str.contains('交付|返修|不良'))&(report['交期判断']=='满足'),'风险等级'] = '一般欠料风险'
            report.loc[(report['类别分类']=='生产关键物料')&(report['当前状态']=='未下单')&(report['超3天未下单']=='超3天未下单'),'风险等级'] = '重大欠料风险'
            report.loc[(report['类别分类']=='生产关键物料')&(report['当前状态']=='未下单')&(report['请购天数']<=3)&(report['请购天数']>=1),'风险等级'] = '较大欠料风险'
            report.loc[(report['类别分类']=='常规物料')&(report['当前状态']=='未下单')&(report['请购天数']<=3)&(report['请购天数']>=1),'风险等级'] = '一般欠料风险'
            report.loc[(report['类别分类']=='常规物料')&(report['当前状态']=='未下单')&(report['超3天未下单']=='超3天未下单'),'风险等级'] = '较大欠料风险'
            report.loc[(report['类别分类']=='生产关键物料')&(report['交期差异天数']<0)&(report['交期差异天数']<-3),'风险等级'] = '重大欠料风险'
            report.loc[(report['类别分类']=='生产关键物料')&(report['交期差异天数']<0)&(report['交期差异天数']<=-1)&(report['交期差异天数']>=-3),'风险等级'] = '较大欠料风险'
            report.loc[(report['类别分类']=='常规物料')&(report['交期差异天数']<0)&(report['交期差异天数']<-3),'风险等级'] = '较大欠料风险'
            report.loc[(report['类别分类']=='常规物料')&(report['交期差异天数']<0)&(report['交期差异天数']<=-1)&(report['交期差异天数']>=-3),'风险等级'] = '一般欠料风险'
            report.loc[(report['类别分类']=='生产关键物料')&(report['交期判断']=='无交期'),'风险等级'] = '重大欠料风险'
            report.loc[(report['类别分类']=='常规物料')&(report['交期判断']=='无交期'),'风险等级'] = '较大欠料风险'

            report['请购日期'] = pd.to_datetime(report["请购日期"],errors='coerce')
            report['采购日期'] = pd.to_datetime(report["采购日期"],errors='coerce')
            report['采购确认日期'] = pd.to_datetime(report["采购确认日期"],errors='coerce')
            report['需求日期'] = pd.to_datetime(report["需求日期"],errors='coerce')
            report['供应商回复交期'] = pd.to_datetime(report["供应商回复交期"],errors='coerce')
            report['数据审核日'] = pd.to_datetime(report["数据审核日"], errors='coerce')
            report['最新需求日期'] = pd.to_datetime(report["最新需求日期"],errors='coerce')
            #report['CEG确认日期'] = pd.to_datetime(report["CEG确认日期"],errors='coerce').dt.strftime('%Y-%m-%d')
            del report['需求日期1']
            report['延期'] = compare_date1(report['请购日期'], device_date)
            report.loc[report['当前状态'].str.contains('暂停|留置'),'延期']=''
            ############判断是否有补充日期表
            screm.insert(INSERT, '\n正在处理需求日期补录表...', '\n')
            window.update()
            path = r'数据源'
            path_file = os.listdir(path)
            for file in path_file:
                if '需求日期补录表' in file:
                    need_time_path = r'数据源\需求日期补录表'
                    # need_price = ['料件编号', '最新价格']
                    need_time_file= os.listdir(need_time_path)
                    for i in need_time_file:
                        if '~$' in i:
                            need_time_file.remove(i)

                    if os.listdir(need_time_path ):
                        for i in range(len(need_time_file)):
                            if str(need_time_file[i]).count('~$') == 0:
                                # report_item = pd.read_excel(filePath1 + '/' + str(file_name1[i]))
                                need_time= pd.read_excel(need_time_path+ '\\' + need_time_file[i])[['项目编号', '模组号','补充需求日期']] #补充需求日期
                        need_time['项目编号'] = need_time['项目编号'] .fillna('')
                        need_time['模组号'] = need_time['模组号'].fillna('')
                        need_time['项目编号'] = need_time['项目编号'].fillna('')
                        need_time['补充需求日期'] = need_time['补充需求日期'].fillna(default_date)
                        need_time['补充需求日期']= pd.to_datetime(need_time['补充需求日期'],errors='coerce')
                        need_time=need_time[(need_time['补充需求日期']!=pd.Timestamp(1990, 1, 1))|(need_time['项目编号']!='')].reset_index(drop=True)
                        need_time1=need_time[need_time['模组号']==''][['项目编号','补充需求日期']].reset_index(drop=True)
                        need_time1=need_time1.drop_duplicates(subset=['项目编号']).reset_index(drop=True)
                        need_time2 = need_time[need_time['模组号']!=''][['项目编号','模组号', '补充需求日期']].reset_index(drop=True)
                        need_time2['库存管理特征']=need_time2['项目编号']+'-'+need_time2['模组号']
                        need_time2=need_time2.drop_duplicates(subset=['库存管理特征']).reset_index(drop=True)
                        ####物料表拉日期
                        report['录入需求日期1']=pd.merge(report,need_time1,on='项目编号',how='left')['补充需求日期']
                        report['录入需求日期1'] = report['录入需求日期1'].fillna(default_date)
                        report['录入需求日期1'] = pd.to_datetime(report['录入需求日期1'], errors='coerce')
                        report['录入需求日期'] = pd.merge(report, need_time2, on='库存管理特征', how='left')['补充需求日期']
                        report['录入需求日期']=report['录入需求日期'].fillna(default_date)
                        report['录入需求日期']=pd.to_datetime(report['录入需求日期'], errors='coerce')
                        report.loc[report['录入需求日期']==pd.Timestamp(1990, 1, 1),'录入需求日期']=report['录入需求日期1']
                        del report['录入需求日期1']
                        ##判断新的交期差异天数
                        report.loc[:, '交期差异天数1'] = subtract3(report['录入需求日期'], report['使用交期'])
                        report.loc[(report['录入需求日期'] >= report['使用交期']) , '交期差异天数1'] = 0
                        report.loc[(report['使用交期'] >= report['录入需求日期']) & (report['录入需求日期'] != pd.Timestamp(1990, 1, 1)) &(datetime.datetime.today().date() > report['使用交期'].dt.date), '交期差异天数1']=(report['录入需求日期'].dt.date - datetime.datetime.today().date()).apply(lambda x: x / np.timedelta64(1, 'D'))
                        report.loc[(report['录入需求日期'] >= report['使用交期']) & (report['录入需求日期'] != pd.Timestamp(1990, 1, 1))&(datetime.datetime.today().date() > report['录入需求日期'].dt.date), '交期差异天数1'] = (report['录入需求日期'].dt.date - datetime.datetime.today().date()).apply(lambda x: x / np.timedelta64(1, 'D'))
                        report.loc[(report['当前状态'] == '已到货') | (report['当前状态'] == '关闭/暂停') | (report['当前状态'] == '已入库')|(report['当前状态'] == '留置') | (report['当前状态'] == '未下单'), '交期差异天数1'] = 0
                        report['交期差异天数1'] = pd.to_numeric(report['交期差异天数1'])
                        ##开始风险判断
                        report['MC风险判断']=''
                        report.loc[(report['类别分类'] == '生产关键物料')&(report['当前状态'].str.contains('未下单')) & (report['设计延期天数']<5), 'MC风险判断'] = '中风险'
                        report.loc[(report['类别分类'] == '生产关键物料')&(report['当前状态'].str.contains('未下单')) & (report['设计延期天数']>=5), 'MC风险判断'] = '高风险'
                        report.loc[(report['类别分类'] == '生产关键物料')&(report['当前状态'].str.contains('未下单')) & (report['请购天数']>=3), 'MC风险判断'] = '高风险'
                        report.loc[(report['类别分类'] == '生产关键物料')&(report['当前状态'].str.contains('未下单')) & (report['请购天数']<3), 'MC风险判断'] = '中风险'
                        report.loc[(report['类别分类'] == '生产关键物料') & (report['当前状态'].str.contains('未下单')==False)&(report['交期差异天数1']>-5), 'MC风险判断'] = '中风险'
                        report.loc[(report['类别分类'] == '生产关键物料') & (report['当前状态'].str.contains('未下单') == False)&(report['交期差异天数1']<=-5), 'MC风险判断'] = '高风险'

                        report.loc[(report['类别分类'] == '常规物料') & (report['当前状态'].str.contains('未下单')) & (report['设计延期天数'] < 5), 'MC风险判断'] = '低风险'
                        report.loc[(report['类别分类'] == '常规物料') & (report['当前状态'].str.contains('未下单')) & (report['设计延期天数'] >= 5), 'MC风险判断'] = '中风险'
                        report.loc[(report['类别分类'] == '常规物料') & (report['当前状态'].str.contains('未下单')) & (report['请购天数'] >= 3), 'MC风险判断'] = '中风险'
                        report.loc[(report['类别分类'] == '常规物料') & (report['当前状态'].str.contains('未下单')) & (report['请购天数'] < 3), 'MC风险判断'] = '低风险'
                        report.loc[(report['类别分类'] == '常规物料') & (report['当前状态'].str.contains('未下单') == False) & (report['交期差异天数1'] > -5), 'MC风险判断'] = '低风险'
                        report.loc[(report['类别分类'] == '常规物料') & (report['当前状态'].str.contains('未下单') == False) & (report['交期差异天数1'] <=-5), 'MC风险判断'] = '中风险'
                        report.loc[(report['当前状态'].str.contains('入库|到货|关闭')),'MC风险判断'] = ''
                        del report['交期差异天数1']
                        report['采购负责人']=report['开单人姓名']
                        report['唯一值']=report['采购单号']+report['采购项次'].astype(str).replace('\.0', '', regex=True)
                        report.loc[report['当前状态'].str.contains('未下单'), '唯一值'] = ''

                        report.loc[report['当前状态'].str.contains('未下单'),'采购负责人']=report['核价采购员姓名']
                        report = report.reindex(columns=['项目编号', '库存管理特征', '作业编号', '模组名称', 'P工序', '系列', '设备','项目阶段',
                                                 '负责人', '请购日期', '请购天数', '来源单号', '来源项次',
                                                 '物料请购负责人', '状态码', '行状态', '料件编号',
                                                 '品名', '规格', '需求数量','品牌', '需求日期', '最新需求日期', '零件类型', '表面处理',
                                                 '采购负责人',
                                                 '采购日期', '数据审核日', '采购单号', '采购项次', '供应商名称', '采购确认日期', '供应商回复交期',
                                                 '已收货量', '已入库量', '仓退数量', '验退数量', '未交量',
                                                 '当前状态', '风险等级', '类别分类','采购组别', '唯一值','交期差异天数', '超3天未下单', '交期判断','已检验数量','入库日期','到货日期',
                                                 '物流信息', '大项目名称','使用交期','延期',
                                                 '采购延期说明','延期进度说明', '是否CEG确认', 'CEG备注', 'CEG确认货期(天)', '设计备注', '交货地址栏(请购单)', '交货地址栏(采购单)','请购单备注','采购确认日期履历', '最新需求日期变更履历','MC负责人','核价采购员姓名','退货快递单号', 'MC风险判断','录入需求日期', '是否大件物料', '设计延期天数','理由码说明','旧版料号','备注说明'])

                        report['录入需求日期'] = pd.to_datetime(report['录入需求日期'], errors='coerce').dt.strftime('%Y-%m-%d')
                        report['录入需求日期'] = ['' if i == '1990-01-01' else i for i in report['录入需求日期']]
                    else:
                        # print("请检查项目对照表，当前无文件")
                        screm.insert(INSERT, '\n当前需求日期补录表无文件，请确认是否需提供', '\n')
                        window.update()
                        report = report.reindex(columns=
                                                ['项目编号', '库存管理特征', '作业编号', '模组名称', 'P工序', '系列', '设备', '项目阶段',
                                                 '负责人', '请购日期', '请购天数', '来源单号', '来源项次',
                                                 '物料请购负责人', '状态码', '行状态', '料件编号',
                                                 '品名', '规格', '需求数量', '品牌', '需求日期', '最新需求日期', '零件类型', '表面处理',
                                                 '采购负责人',
                                                 '采购日期', '数据审核日', '采购单号', '采购项次', '供应商名称', '采购确认日期', '供应商回复交期',
                                                 '已收货量', '已入库量', '仓退数量', '验退数量', '未交量',
                                                 '当前状态', '风险等级', '类别分类', '采购组别', '唯一值', '交期差异天数', '超3天未下单', '交期判断', '已检验数量',
                                                 '入库日期', '到货日期',
                                                 '物流信息', '大项目名称', '使用交期', '延期',
                                                 '采购延期说明','延期进度说明', '是否CEG确认', 'CEG备注', 'CEG确认货期(天)', '设计备注', '交货地址栏(请购单)', '交货地址栏(采购单)',
                                                 '请购单备注', '采购确认日期履历', '最新需求日期变更履历', 'MC负责人', '核价采购员姓名', '退货快递单号', 'MC风险判断',
                                                 '录入需求日期', '是否大件物料', '设计延期天数','理由码说明','旧版料号','备注说明'])

                        report=report.fillna('')
            if '需求日期补录表' not in path_file:
                screm.insert(INSERT, '\n当前无需求日期补录表文件，请确认是否需提供', '\n')
                window.update()
                report = report.reindex(columns=
                                        ['项目编号', '库存管理特征', '作业编号', '模组名称', 'P工序', '系列',
                                         '设备', '项目阶段',
                                         '负责人', '请购日期', '请购天数', '来源单号', '来源项次',
                                         '物料请购负责人', '状态码', '行状态', '料件编号',
                                         '品名', '规格', '需求数量', '品牌', '需求日期', '最新需求日期', '零件类型',
                                         '表面处理',
                                         '采购负责人',
                                         '采购日期', '数据审核日', '采购单号', '采购项次', '供应商名称',
                                         '采购确认日期', '供应商回复交期',
                                         '已收货量', '已入库量', '仓退数量', '验退数量', '未交量',
                                         '当前状态', '风险等级', '类别分类', '采购组别', '唯一值', '交期差异天数',
                                         '超3天未下单', '交期判断', '已检验数量',
                                         '入库日期', '到货日期',
                                         '物流信息', '大项目名称', '使用交期', '延期',
                                         '采购延期说明','延期进度说明', '是否CEG确认', 'CEG备注', 'CEG确认货期(天)', '设计备注',
                                         '交货地址栏(请购单)', '交货地址栏(采购单)',
                                         '请购单备注', '采购确认日期履历', '最新需求日期变更履历', 'MC负责人',
                                         '核价采购员姓名', '退货快递单号', 'MC风险判断',
                                         '录入需求日期', '是否大件物料', '设计延期天数', '理由码说明','旧版料号','备注说明'])
                    ###采购组别
            ############判断是否有补充日期表
            screm.insert(INSERT, '\n正在处理采购组别表...', '\n')
            window.update()
            buyer_path = r'数据源\采购组别'
            # need_price = ['料件编号', '最新价格']
            buyer_file = os.listdir(buyer_path)
            for i in buyer_file:
                if '~$' in i:
                    buyer_file.remove(i)
            if os.listdir(buyer_path):
                for i in range(len(buyer_file)):
                    if str(buyer_file[i]).count('~$') == 0 and 'xlsx' in str(buyer_file[i]):
                        # report_item = pd.read_excel(filePath1 + '/' + str(file_name1[i]))
                        buyer = pd.read_excel(buyer_path + '\\' + buyer_file[i])[['姓名', '组别']]  # 补充需求日期
                buyer = buyer.drop_duplicates(subset=['姓名']).reset_index(drop=True)
                buyer['姓名'] = buyer['姓名'].fillna('').astype(str)
                buyer['组别'] = buyer['组别'].fillna('').astype(str)
                report['采购负责人'] = report['采购负责人'].fillna('')
                report['采购组别'] = pd.merge(report, buyer, left_on=['采购负责人'], right_on=['姓名'], how='left')[
                    '组别']
                report = report.fillna('')
            else:
                screm.insert(INSERT, '\n无采购组别表...', '\n')
                window.update()
            screm.insert(INSERT, '\n正在处理系统未抛单项次...', '\n')
            window.update()
            apsp_path = r'数据源\\apsp600'
            # need_price = ['料件编号', '最新价格']
            apsp_file = os.listdir(apsp_path)
            for i in apsp_file:
                if '~$' in i:
                    apsp_file.remove(i)
            if os.listdir(apsp_path):
                for i in range(len(apsp_file)):
                    if str(apsp_file[i]).count('~$') == 0 :
                        # report_item = pd.read_excel(filePath1 + '/' + str(file_name1[i]))
                        apsp = pd.read_excel(apsp_path + '\\' + apsp_file[i])[['项目编号']]  # 补充需求日期

                apsp['项目编号']=apsp['项目编号'].fillna('').astype(str)
                apsp['计数']='计数'
                apsp_use=apsp.groupby(['项目编号']).agg({'计数': 'count'}).add_suffix('').reset_index()
                apsp_use['库存管理特征']=apsp['项目编号'].astype(str)+'-无模组号'

            else:
                screm.insert(INSERT, '\n无统未抛单表...', '\n')
                window.update()
                apsp_use=pd.DataFrame(columns=['项目编号','库存管理特征','计数'])

            time3 = time.time()
            print('%s物控表列处理耗时:%d秒' % (bookname,time3-time2))
            print('已完成{:.0%}'.format(0.6))
            screm.insert(INSERT, '\n%s物控表列处理耗时:%d秒' % (bookname,time3-time2))
            screm.insert(INSERT, '\n已完成{:.0%}'.format(0.6))
            window.update()
            # 处理整体数据汇总表的列数据
            item_report_need=['项数', '项目号', '系列', '设备名称',   #'项目负责人',
                     '需求齐料时间','设备数量',
                     '计划出货日期', '计划下单总项数', '设计下单完成日期']
            for col in item_report.columns:
                if col not in item_report_need:
                    del item_report[col]
            #到料90%交期——项目
            product90=pd.DataFrame(report[(report['当前状态']!='关闭/暂停')&(report['当前状态']!='留置')].groupby(['项目编号','使用交期'])['料件编号'].count()).add_suffix('-统计').sort_values(by=['项目编号','使用交期']).reset_index()
            product90.loc[(product90['使用交期']=='1990-01-01'),'料件编号-统计'] = 0
            product90['累计到料统计']=''
            product90['总料数']=''
            product90['累计到料统计占比']=''
            module1=pd.DataFrame(report[(report['当前状态']!='关闭/暂停')&(report['当前状态']!='留置')].groupby(['项目编号'])['料件编号'].count()).add_suffix('-统计之和').reset_index()
            product90['总料数']=pd.merge(product90,module1,how='left',left_on='项目编号',right_on='项目编号')['料件编号-统计之和']
            if len( product90)!=0:
                product90['累计到料统计'][0]=product90['料件编号-统计'][0]
            for  i in range(1,len(product90)):
                if product90['项目编号'][i]==product90['项目编号'][i-1]:
                    product90['累计到料统计'][i]=product90['累计到料统计'][i-1]+product90['料件编号-统计'][i]
                if product90['项目编号'][i]!=product90['项目编号'][i-1]:
                    product90['累计到料统计'][i]=product90['料件编号-统计'][i]
            product90['累计到料统计占比']=product90['累计到料统计']/product90['总料数']

            product90['累计到料统计占比']=pd.to_numeric(product90['累计到料统计占比'])
            product90_1=product90[product90['累计到料统计占比']>=0.95].reset_index(drop=True)
            #product90['累计到料统计占比']= product90['累计到料统计占比'].apply(lambda x:format(x,'.2%'))
            product90_2=pd.DataFrame(product90_1.groupby(['项目编号'])['使用交期','累计到料统计占比'].min()).reset_index()

            item_report['到料95%交期']=pd.merge(item_report, product90_2, left_on ='项目号',right_on='项目编号', how ='left')['使用交期']
            #002
            report_new=pd.concat([report,report002[['项目编号','库存管理特征','料件编号']]], ignore_index=True).reset_index(drop=True)
            project_transplay=pd.DataFrame(report_new.groupby(['项目编号','库存管理特征'])['料件编号'].count()).add_suffix('统计').reset_index()
            project_transplay=pd.DataFrame(project_transplay.groupby(['项目编号'])['料件编号统计'].sum()).add_suffix('-之和').reset_index()
            item_report['累计下单总项数']=pd.merge(item_report,project_transplay,left_on='项目号',right_on='项目编号',how='left')['料件编号统计-之和']
            item_report['系统未抛单项数'] = \
            pd.merge(item_report, apsp_use, left_on='项目号', right_on='项目编号', how='left')['计数'].fillna(0)

            #200
            report200['库存管理特征'] = report200['项目号'] + '-' + report200['作业编号']
            report200drop=pd.DataFrame(report200.groupby(['项目号','直接上阶料号'])['元件料号'].count()).add_suffix('统计').reset_index()
            report200drop1=pd.DataFrame(report200drop.groupby(['项目号'])['元件料号统计'].sum()).add_suffix('-之和').reset_index()
            item_report['下单总项数']=pd.merge(item_report,report200drop1,left_on='项目号',right_on='项目号',how='left')['元件料号统计-之和']

            plantime_delta = list(datetime.datetime.today() - item_report['设计下单完成日期'])
            item_report.loc[:, '超计划下单天数'] = [item.days if item.days < 1000 else 0 for item in plantime_delta]

            #加入新表
            #需求日期需更新为最新需求时间
            unique_feature=report[['项目编号','库存管理特征','需求日期','系列']].drop_duplicates(subset=['库存管理特征']).rename(columns={'库存管理特征':'项目号','需求日期':'需求齐料时间'})
            #####如无模组号的不在这个表里就插入一条
            it = list(item_report.drop_duplicates(subset=['项目号']).reset_index(drop=True)['项目号'])
            for it1 in it:
                it2 = str(it1 + '-无模组号')
                series=item_report[item_report['项目号'].str.contains(it1)].reset_index(drop=True)['系列'][0]
                if len(unique_feature[unique_feature['项目号'].str.contains(it2)]) < 1 and len(project_transplay[project_transplay['项目编号']==it1])>0  :
                    new_row = pd.DataFrame({'项目编号': [it1], '项目号': [it1 + '-无模组号'], '需求日期': [default_date], '系列': [series]}, index=[len(unique_feature)])
                    unique_feature= pd.concat([unique_feature, new_row], ignore_index=True).reset_index(drop=True)

            unique_feature=unique_feature.reindex(columns=
                    ['项数', '项目号', '系列', '设备名称', #'项目负责人',
                     '需求齐料时间','设备数量','计划出货日期','参考项目号',
                      '总项数',"D1         [KCL物料]","D2             [标准件]","D3             [非标件]","D4             [现场模组]",
                     '计划下单总项数',"D1         (KCL物料)","D2             (标准件)","D3             (非标件)","D4             (现场模组)",'设计下单完成日期',
                     '下单总项数',"D1             <KCL物料>","D2                <标准件>","D3                <非标件>","D4                <现场模组>",
                     '累计下单总项数',
                     '实际下单总项数','超计划日期下单总项数', '超计划日期下单比例', '实际下单进度','超计划下单天数','系统未抛单项数','抛单进度%',
                     '采购已下单总项数'  #, '超5天下单总量'
                        , '采购下单率', '交期不满足需求日期',
                     '采购未下单总项数', '超3天未下单总量', '超3天未下单比例',
                      '已到货项数', '未到货项数',
                      '延期项数', '退回返修项数','关闭/暂停项数','形态转换项数','已入库项数',
                    '到料率','入库率','项目编号','到料95%交期','留置项数','实际到料率（考虑设计下单进度）','实际入库率（考虑设计下单进度）'])
            unique_feature['需求齐料时间']=unique_feature['需求齐料时间'].fillna(default_date)
            unique_feature=unique_feature.reset_index(drop=True)
            module_time=report[report['最新需求日期']!=pd.Timestamp(1990, 1, 1)][['库存管理特征','最新需求日期']].drop_duplicates(subset=['库存管理特征']).reset_index(drop=True)
            unique_merge=pd.merge(unique_feature,module_time,how='left',left_on='项目号',right_on='库存管理特征').reset_index(drop=True)
            unique_merge['最新需求日期']=unique_merge['最新需求日期'].fillna(default_date)
            unique_merge['最新需求日期'] = pd.to_datetime(unique_merge['最新需求日期'], errors='coerce')
            unique_merge.loc[(unique_merge['最新需求日期'] !=pd.Timestamp(1990, 1, 1)), '需求齐料时间'] = unique_merge['最新需求日期']
            unique_feature['需求齐料时间']=unique_merge['需求齐料时间']
            #到料90%交期
            module90=pd.DataFrame(report[(report['当前状态']!='关闭/暂停')&(report['当前状态']!='留置')].groupby(['库存管理特征','使用交期'])['料件编号'].count()).add_suffix('-统计').sort_values(by=['库存管理特征','使用交期']).reset_index()
            module90.loc[(module90['使用交期']=='1990-01-01'),'料件编号-统计'] = 0
            module90['累计到料统计']=''
            module90['总料数']=''
            module90['累计到料统计占比']=''
            module1=pd.DataFrame(report[(report['当前状态']!='关闭/暂停')&(report['当前状态']!='留置')].groupby(['库存管理特征'])['料件编号'].count()).add_suffix('-统计之和').reset_index()
            module90['总料数']=pd.merge(module90,module1,how='left',left_on='库存管理特征',right_on='库存管理特征')['料件编号-统计之和']
            if len(module90)!=0:
                module90['累计到料统计'][0]=module90['料件编号-统计'][0]
            for  i in range(1,len(module90)):
                if module90['库存管理特征'][i]==module90['库存管理特征'][i-1]:
                    module90['累计到料统计'][i]=module90['累计到料统计'][i-1]+module90['料件编号-统计'][i]
                if module90['库存管理特征'][i]!=module90['库存管理特征'][i-1]:
                    module90['累计到料统计'][i]=module90['料件编号-统计'][i]
            module90['累计到料统计占比']=module90['累计到料统计']/module90['总料数']

            module90['累计到料统计占比']=pd.to_numeric(module90['累计到料统计占比'])
            module90_1=module90[module90['累计到料统计占比']>=0.95].reset_index(drop=True)
            #module90['累计到料统计占比']= module90['累计到料统计占比'].apply(lambda x:format(x,'.2%'))
            module90_2=pd.DataFrame(module90_1.groupby(['库存管理特征'])['使用交期','累计到料统计占比'].min()).reset_index()

            unique_feature=unique_feature.reset_index(drop=True)
            unique_feature['到料95%交期']=pd.merge(unique_feature, module90_2, left_on ='项目号',right_on='库存管理特征', how ='left')['使用交期']


            feauture_drop1=pd.DataFrame(report.groupby(['库存管理特征'])['料件编号'].count()).add_suffix('统计').reset_index()
            unique_feature=unique_feature.reset_index(drop=True)
            unique_feature['实际下单总项数-下单']=pd.merge(unique_feature, feauture_drop1, left_on ='项目号',right_on='库存管理特征', how ='left')['料件编号统计'].fillna(0)

            design_drop1=pd.DataFrame(report[report['延期']=='是'].groupby(['库存管理特征'])['料件编号'].count()).add_suffix('统计').reset_index()
            del report['延期']
            unique_feature=unique_feature.reset_index(drop=True)
            unique_feature['超计划日期下单总项数']=pd.merge(unique_feature, design_drop1, left_on ='项目号',right_on='库存管理特征', how ='left')['料件编号统计']

            noorder_drop1=pd.DataFrame(report[(report['当前状态']=='未下单')].groupby(['库存管理特征','当前状态'])['料件编号'].count()).add_suffix('统计').reset_index()
            unique_feature=unique_feature.reset_index(drop=True)
            unique_feature['采购未下单总项数']=pd.merge(unique_feature, noorder_drop1, left_on ='项目号',right_on='库存管理特征', how ='left')['料件编号统计']

            nofit_drop1=pd.DataFrame(report[(report['交期判断']=='不满足')&((report['当前状态']=='采购中')|(report['当前状态']=='交付延误')|(report['当前状态']=='制程返修')|(report['当前状态']=='来料不良'))].groupby(['库存管理特征'])['料件编号'].count()).add_suffix('统计').reset_index()
            unique_feature=unique_feature.reset_index(drop=True)
            unique_feature['交期不满足需求日期']=pd.merge(unique_feature, nofit_drop1, left_on ='项目号',right_on='库存管理特征', how ='left')['料件编号统计']

            over_five=pd.DataFrame(report[(report['超3天未下单']=='超3天未下单')].groupby(['库存管理特征','超3天未下单'])['料件编号'].count()).add_suffix('统计').reset_index()
            unique_feature=unique_feature.reset_index(drop=True)
            unique_feature['超3天未下单总量']=pd.merge(unique_feature, over_five, left_on ='项目号',right_on='库存管理特征', how ='left')['料件编号统计']

            in_drop1=pd.DataFrame(report[(report['当前状态']=='已入库')].groupby(['库存管理特征','当前状态'])['料件编号'].count()).add_suffix('统计').reset_index()
            unique_feature=unique_feature.reset_index(drop=True)
            unique_feature['已入库项数']=pd.merge(unique_feature, in_drop1, left_on ='项目号',right_on='库存管理特征', how ='left')['料件编号统计']

            get_drop1=pd.DataFrame(report[(report['当前状态']=='已到货')].groupby(['库存管理特征'])['料件编号'].count()).add_suffix('统计').reset_index()
            get_drop1=get_drop1.groupby(by=['库存管理特征'])['料件编号统计'].sum().reset_index()
            unique_feature=unique_feature.reset_index(drop=True)
            unique_feature['已到货项数']=pd.merge(unique_feature, get_drop1, left_on ='项目号',right_on='库存管理特征', how ='left')['料件编号统计']

            close_drop1=pd.DataFrame(report[(report['当前状态']=='关闭/暂停')].groupby(['库存管理特征','当前状态'])['料件编号'].count()).add_suffix('统计').reset_index()
            unique_feature=unique_feature.reset_index(drop=True)
            unique_feature['关闭/暂停项数']=pd.merge(unique_feature, close_drop1, left_on ='项目号',right_on='库存管理特征', how ='left')['料件编号统计']

            remain_drop1=pd.DataFrame(report[(report['当前状态']=='留置')].groupby(['库存管理特征','当前状态'])['料件编号'].count()).add_suffix('统计').reset_index()
            unique_feature=unique_feature.reset_index(drop=True)
            unique_feature['留置项数']=pd.merge(unique_feature, remain_drop1, left_on ='项目号',right_on='库存管理特征', how ='left')['料件编号统计']

            delay_drop1=pd.DataFrame(report[(report['当前状态']=='交付延误')].groupby(['库存管理特征','当前状态'])['料件编号'].count()).add_suffix('统计').reset_index()
            unique_feature=unique_feature.reset_index(drop=True)
            unique_feature['延期项数']=pd.merge(unique_feature, delay_drop1, left_on ='项目号',right_on='库存管理特征', how ='left')['料件编号统计']

            back_drop1=pd.DataFrame(report[(report['当前状态']=='制程返修')|(report['当前状态']=='来料不良')].groupby(['库存管理特征','当前状态'])['料件编号'].count()).add_suffix('统计').reset_index()
            unique_feature=unique_feature.reset_index(drop=True)
            unique_feature['退回返修项数']=pd.merge(unique_feature, back_drop1, left_on ='项目号',right_on='库存管理特征', how ='left')['料件编号统计']

            noget_drop1=pd.DataFrame(report[(report['当前状态']=='未下单')|(report['当前状态']=='交付延误')|(report['当前状态']=='采购中')|(report['当前状态']=='制程返修')|(report['当前状态']=='来料不良')].groupby(['库存管理特征'])['料件编号'].count()).add_suffix('统计').reset_index()
            unique_feature=unique_feature.reset_index(drop=True)
            unique_feature['未到货项数']=pd.merge(unique_feature, noget_drop1, left_on ='项目号',right_on='库存管理特征', how ='left')['料件编号统计']
            unique_feature=unique_feature.reset_index(drop=True)

            orderd_drop1=pd.DataFrame(report[(report['当前状态']!='未下单')&(~report['当前状态'].str.contains('关闭|留置'))].groupby(['库存管理特征'])['料件编号'].count()).add_suffix('统计').reset_index()
            unique_feature['采购已下单总项数']=pd.merge(unique_feature, orderd_drop1, left_on ='项目号',right_on='库存管理特征', how ='left')['料件编号统计']
            unique_feature=unique_feature.reset_index(drop=True)

            drop1_200=pd.DataFrame(report200.groupby(['库存管理特征'])['元件料号'].count()).add_suffix('统计').reset_index()
            unique_feature['下单总项数']=pd.merge(unique_feature,drop1_200,left_on ='项目号',right_on='库存管理特征', how ='left')['元件料号统计']
            unique_feature['下单总项数'].fillna(0)
            unique_feature.loc[(pd.isnull(unique_feature['下单总项数'])==True),'下单总项数'] = 0

            ##transplay1=pd.DataFrame(report_new.groupby(['库存管理特征'])['料件编号'].nunique()).add_suffix('去重统计').reset_index()
            transplay1 = pd.DataFrame(report_new.groupby(['库存管理特征'])['料件编号'].count()).add_suffix('统计').reset_index()
            unique_feature=unique_feature.reset_index(drop=True)
            unique_feature['累计下单总项数']=pd.merge(unique_feature, transplay1, left_on ='项目号',right_on='库存管理特征', how ='left')['料件编号统计']
            unique_feature['累计下单总项数']=unique_feature['累计下单总项数'].fillna(0)
            #unique_feature.loc[(pd.isnull(unique_feature['累计下单总项数'])==True),'累计下单总项数'] = ''
            unique_feature['系统未抛单项数']=pd.merge(unique_feature,apsp_use,left_on='项目号',right_on='库存管理特征',how='left')['计数'].fillna(0)


            feature_int=['实际下单总项数','超计划日期下单总项数','超计划下单天数','系统未抛单项数','采购未下单总项数','交期不满足需求日期'
                         ,'超3天未下单总量','形态转换项数','已入库项数','已到货项数','关闭/暂停项数','采购已下单总项数'
                        ,'未到货项数','延期项数','退回返修项数','下单总项数','留置项数','实际下单总项数-下单']
            unique_feature[feature_int]=unique_feature[feature_int].fillna(0)
            unique_feature['实际下单总项数'] = unique_feature['累计下单总项数'] - unique_feature['关闭/暂停项数'] - \
                                               unique_feature['留置项数']

            # 做项目总计行
            # 做项目总计行
            statics_sum_col = [ '实际下单总项数',
             '超计划日期下单总项数', '采购已下单总项数', #'超5天下单总量',
             '交期不满足需求日期', '采购未下单总项数', '超3天未下单总量',  '已到货项数', '未到货项数',
                      '延期项数', '退回返修项数','关闭/暂停项数','形态转换项数','已入库项数','留置项数','实际下单总项数-下单']
            unique_feature1=unique_feature.copy()


            #直接从report获取，不提取 unique_feature1['项目编号']=unique_feature1['项目号'].str.split('-', expand=True)[0]
            item_report = item_report.reset_index(drop=True)
            unique_feature_group = unique_feature1.groupby(unique_feature1['项目编号'])[statics_sum_col].sum()
            unique_feature_group = unique_feature_group.reset_index()
            item_report=item_report.reset_index()
            item_report=pd.merge(item_report,unique_feature_group,left_on='项目号',right_on='项目编号',how='left')
            del unique_feature['项目编号']
            item_report = item_report.reindex(columns=
                    ['项数', '项目号', '系列', '设备名称',         #'项目负责人',
                     '需求齐料时间','设备数量','计划出货日期','参考项目号',
                       '总项数',"D1         [KCL物料]","D2             [标准件]","D3             [非标件]","D4             [现场模组]",
                     '计划下单总项数',"D1         (KCL物料)","D2             (标准件)","D3             (非标件)","D4             (现场模组)",'设计下单完成日期',
                     '下单总项数',"D1             <KCL物料>","D2                <标准件>","D3                <非标件>","D4                <现场模组>",
                     '累计下单总项数',
                     '实际下单总项数','超计划日期下单总项数', '超计划日期下单比例', '实际下单进度','超计划下单天数','系统未抛单项数',"抛单进度%",
                     '采购已下单总项数'   #'超5天下单总量'
                        , '采购下单率', '交期不满足需求日期',
                     '采购未下单总项数', '超3天未下单总量', '超3天未下单比例',
                     '已到货项数', '未到货项数',
                      '延期项数', '退回返修项数','关闭/暂停项数','形态转换项数','已入库项数','到料率',
                    '入库率','到料95%交期','留置项数','实际下单总项数-下单','实际到料率（考虑设计下单进度）','实际入库率（考虑设计下单进度）'])
            # 做系列总计行
            statics_sum_col = ['设备数量', '计划下单总项数','下单总项数','累计下单总项数','实际下单总项数',
             '超计划日期下单总项数','系统未抛单项数', '采购已下单总项数', #'超5天下单总量',
             '交期不满足需求日期', '采购未下单总项数', '超3天未下单总量',  '已到货项数', '未到货项数',
                      '延期项数', '退回返修项数','关闭/暂停项数','形态转换项数','已入库项数','留置项数','到料率',
                    '入库率','实际下单总项数-下单']
            item_report_group = item_report.groupby(item_report['系列'].str[:2] + '龠系列总计行')[statics_sum_col].sum()
            item_report_group = item_report_group.reset_index()
            item_report=item_report.reset_index(drop=True)

            item_report = pd.concat([item_report, item_report_group], ignore_index=True)

            item_report = item_report.reset_index(drop=True)
            # item_report['项数'] = item_report['项数'].fillna('Total')
            item_report = item_report.fillna(' ')
            item_report.sort_values(by=['系列','项目号'],ascending=[True,True],inplace=True)
            item_report = item_report.reset_index(drop=True)
            #纵向合并模组和项目

            item_report=pd.concat([item_report,unique_feature]).reset_index(drop=True)
            item_report = item_report.fillna(' ')
            #item_report = item_report.reset_index(drop=True)
            item_report.sort_values(by=['系列','项目号'],inplace=True)
            item_report['系列']=item_report['系列'].replace('龠','',regex=True).astype(str)
            item_report = item_report.reset_index(drop=True)

            #del item_report['项目编号']
            item_report: object=item_report.rename(columns={'项数':'序号' #,'需求齐料时间':'物料需求时间'
            })
            #序号重新排
            item_report['序号']=''
            item_proj=item_report[(item_report['项目号'].str.contains('-...') == False)&(item_report['系列'].str.contains('总计') == False)].reset_index(drop=True)
            j=1
            item_proj['序号'][0]=1
            for   i in range(1,len(item_proj)):
                if str(item_proj['系列'].at[i])[:2]==str(item_proj['系列'].at[i-1])[:2]:
                    j=j+1
                    item_proj['序号'].at[i]=j
                else:
                    j=1
                    item_proj['序号'].at[i]=j
            #workbook1_1[序号]=

            item_report['序号']=pd.merge(item_report,item_proj,on='项目号',how='left')['序号_y']
            item_report['序号']=item_report['序号'].fillna('0')
            item_report['序号']=pd.to_numeric(item_report['序号'], errors='ignore')
            for i in item_report.index:
                if item_report['序号'][i]==0:
                    item_report['序号'][i]=str(item_report['序号'][i-1])

            item_model=item_report[item_report['项目号'].str.contains('-...') == True].reset_index(drop=True)
            m=1
            item_model['序']=item_model['序号']
            if len(item_model)!=0:
                item_model['序号'][0]=str(item_model['序号'][0])+'-1'
            for i in range(1,len(item_model)):
                if (str(item_model['序号'][i])==str(item_model['序'][i-1]))&(item_model['系列'][i]==str(item_model['系列'][i-1])):
                    m=m+1
                    item_model['序号'][i]=str(item_model['序'][i-1])+'-'+str(m)
                else :
                    m=1
                    item_model['序号'][i]=str(item_model['序'][i])+'-'+str(m)
            a=pd.merge(item_report,item_model,on='项目号',how='left')
            a['序号_y']=a['序号_y'].fillna('空')
            a.loc[(a['序号_y']!='空'),'序号_x'] = a['序号_y']
            item_report['序号']=a['序号_x']
            item_report.loc[((item_report['系列'].str.contains('总计') == True)),'序号'] = ''

            item_report_length = len(item_report)
            intcol=['下单总项数','累计下单总项数','实际下单总项数','超计划日期下单总项数','超计划下单天数','系统未抛单项数','采购已下单总项数'
                    ,'交期不满足需求日期','采购未下单总项数','超3天未下单总量','已到货项数'
                    ,'未到货项数','延期项数','退回返修项数','关闭/暂停项数','形态转换项数','已入库项数','留置项数','实际下单总项数-下单']
            item_report[intcol] = item_report[intcol].fillna(0)

            item_report['累计下单总项数']=item_report['累计下单总项数'].fillna(0)
            item_report['下单总项数'] = pd.to_numeric(item_report['下单总项数'],errors='coerce')
            item_report['累计下单总项数'] = pd.to_numeric(item_report['累计下单总项数'],errors='coerce')
            item_report['实际下单总项数'] = pd.to_numeric(item_report['实际下单总项数'],errors='coerce')
            item_report['实际下单总项数-下单'] = pd.to_numeric(item_report['实际下单总项数-下单'], errors='coerce')
            item_report['形态转换项数'] = item_report['累计下单总项数'] - item_report['实际下单总项数-下单']
            del  item_report['实际下单总项数-下单']
            item_report.loc[item_report['形态转换项数']<=0,'形态转换项数']=0
            item_report['计划下单总项数'] = pd.to_numeric(item_report['计划下单总项数'],errors='coerce')
            item_report['采购已下单总项数'] = pd.to_numeric(item_report['采购已下单总项数'],errors='coerce')
            item_report['超3天未下单总量'] = pd.to_numeric(item_report['超3天未下单总量'],errors='coerce')
            item_report['采购未下单总项数'] = pd.to_numeric(item_report['采购未下单总项数'],errors='coerce')
            item_report['形态转换项数'] = pd.to_numeric(item_report['形态转换项数'], errors='coerce')
            item_report['已入库项数'] = pd.to_numeric(item_report['已入库项数'],errors='coerce')
            item_report['超计划日期下单总项数'] = pd.to_numeric(item_report['超计划日期下单总项数'],errors='coerce')
            item_report['已到货项数'] = pd.to_numeric(item_report['已到货项数'],errors='coerce')
            item_report['未到货项数'] = pd.to_numeric(item_report['未到货项数'],errors='coerce')
            item_report['延期项数'] = pd.to_numeric(item_report['延期项数'],errors='coerce')
            item_report['退回返修项数'] = pd.to_numeric(item_report['退回返修项数'],errors='coerce')
            item_report['交期不满足需求日期'] = pd.to_numeric(item_report['交期不满足需求日期'],errors='coerce')
            item_report['关闭/暂停项数'] = pd.to_numeric(item_report['关闭/暂停项数'],errors='coerce')
            item_report['留置项数'] = pd.to_numeric(item_report['留置项数'],errors='coerce')
            item_report['系统未抛单项数'] = pd.to_numeric(item_report['系统未抛单项数'], errors='coerce')
            #item_report.loc[(pd.isnull(item_report['计划下单总项数'])==True),'计划下单总项数'] = 0
            item_report.loc[(pd.isnull(item_report['下单总项数'])==True),'下单总项数'] = 0
            #item_report.loc[(pd.isnull(item_report['已请购+库存转换总项数'])==True),'已请购+库存转换总项数'] = 0
            item_report.loc[(pd.isnull(item_report['实际下单总项数'])==True),'实际下单总项数'] = 0
            item_report.loc[(pd.isnull(item_report['超3天未下单总量'])==True),'超3天未下单总量'] = 0
            item_report.loc[(pd.isnull(item_report['采购已下单总项数'])==True),'采购已下单总项数'] = 0
            item_report.loc[(pd.isnull(item_report['形态转换项数']) == True), '形态转换项数'] = 0
            item_report.loc[(pd.isnull(item_report['已入库项数'])==True),'已入库项数'] = 0
            item_report.loc[(pd.isnull(item_report['超计划日期下单总项数'])==True),'超计划日期下单总项数'] = 0
            item_report.loc[(pd.isnull(item_report['系统未抛单项数']) == True), '系统未抛单项数'] = 0

            item_report.loc[(pd.isnull(item_report['交期不满足需求日期'])==True),'交期不满足需求日期'] = 0
            item_report.loc[(pd.isnull(item_report['采购未下单总项数'])==True),'采购未下单总项数'] = 0
            item_report.loc[(pd.isnull(item_report['已到货项数'])==True),'已到货项数'] = 0
            item_report.loc[(pd.isnull(item_report['关闭/暂停项数'])==True),'关闭/暂停项数'] = 0
            item_report.loc[(pd.isnull(item_report['未到货项数'])==True),'未到货项数'] = 0
            item_report.loc[(pd.isnull(item_report['延期项数'])==True),'延期项数'] = 0
            item_report.loc[(pd.isnull(item_report['退回返修项数'])==True),'退回返修项数'] = 0
            item_report.loc[(pd.isnull(item_report['留置项数'])==True),'退回返修项数'] = 0

            item_report['超计划日期下单比例'] = [
                AdivB_percent(item_report['超计划日期下单总项数'][i],
                              item_report['实际下单总项数'][i])
                for i in range(item_report_length)]
            item_report.loc[(pd.isnull(item_report['累计下单总项数'])==True),'累计下单总项数'] = ''
            item_report.loc[(pd.isnull(item_report['计划下单总项数'])==True),'计划下单总项数'] = ''
            item_report.loc[(item_report['实际下单总项数']==''),'实际下单总项数'] = 0
            item_report['实际下单总项数扣除形态转换']=item_report['实际下单总项数']-item_report['形态转换项数']
            item_report['实际下单进度'] = [
                AdivB_percent(item_report['实际下单总项数'][i],
                              item_report['计划下单总项数'][i])
                for i in range(item_report_length)]

            item_report["抛单进度%"]=[
                AdivB_percent_other(item_report['系统未抛单项数'][i],
                              item_report['下单总项数'][i])
                for i in range(item_report_length)]

            item_report['采购下单率'] = [
                AdivB_percent(item_report['采购已下单总项数'][i],
                              item_report['实际下单总项数扣除形态转换'][i])
                for i in range(item_report_length)]
            del item_report['实际下单总项数扣除形态转换']
            item_report['超3天未下单比例'] = [
                AdivB_percent(item_report['超3天未下单总量'][i],
                              item_report['采购未下单总项数'][i])
                for i in range(item_report_length)]
            item_report['入库率']=(item_report['已入库项数']+item_report['形态转换项数'])/(item_report['实际下单总项数'])

            item_report['到料率'] = (item_report['已到货项数'] +item_report['已入库项数']+item_report['形态转换项数'])/ (
                        item_report['实际下单总项数'])

            item_report['实际到料率（考虑设计下单进度）']=[
                sum_cal(item_report['到料率'][i],
                        item_report['实际下单进度'][i])
                for i in range(item_report_length)]#, '实际到料率（考虑设计下单进度）'
            item_report['实际到料率（考虑设计下单进度）']=item_report['实际到料率（考虑设计下单进度）'].fillna('')

            item_report['实际入库率（考虑设计下单进度）'] = [
                sum_cal(item_report['入库率'][i],
                    item_report['实际下单进度'][i])
                for i in range(item_report_length)]
            item_report['实际入库率（考虑设计下单进度）'] = item_report['实际入库率（考虑设计下单进度）'].fillna('')
            #item_report['实际入库率（考虑设计下单进度）'] = item_report['入库率'] + item_report['实际下单进度']

            item_report.loc[(item_report['实际下单总项数'] == 0), '入库率'] = ''
            item_report['入库率'] = item_report['入库率'].fillna('')
            item_report.loc[(item_report['实际下单总项数'] == 0), '到料率'] = ''


            item_report['到料率'] = item_report['到料率'].fillna('')
            # del item_report['留置项数'] #后面再删
            # df_to_xlsx(objpath=r'物料管控表.xlsx',
            #                     df_to_sheet_seq=[(report, '物料管控表'),
            #                             (item_report, 'XXXXX项目物料整体数据一览表')])
            time4 = time.time()
            print('%s整体数据汇总列处理耗时:%d秒' % (bookname,time4-time3))
            print('已完成{:.0%}'.format(0.8))
            screm.insert(INSERT, '\n%s整体数据汇总列处理耗时:%d秒' % (bookname,time4-time3))
            screm.insert(INSERT, '\n已完成{:.0%}'.format(0.8))
            window.update()
            ######################comp012在途占用公共仓明细表
            cpmq012_use=cpmq012[cpmq012['项目号'].isin(item_report['项目号'])].reset_index(drop=True)
            cpmq012_use['在途占用未入库数量']=cpmq012_use['在途占用数量']-cpmq012_use['在途占用已入库数量']

            # 创建xlsx，写入数据并设置格式
            now_time = time.strftime("%Y-%m-%d-%H",time.localtime(time.time()))
            book_name='物料管控表'+bookname+now_time
            workbook = xlsxwriter.Workbook(book_name+'.xlsx', {'nan_inf_to_errors': True})
            worksheet2 = workbook.add_worksheet('整体数据汇总') # 创建sheet
            worksheet1 = workbook.add_worksheet('物料管控表')
            worksheet1_3 = workbook.add_worksheet('BOM比对结果')
            worksheet3= workbook.add_worksheet('形态转换明细表')
            worksheet4= workbook.add_worksheet('项目到料趋势')
            worksheet5 = workbook.add_worksheet('在途占用公共仓明细表')

            title_format = workbook.add_format({'font_name': 'Arial',
                                                'font_size': 14,
                                                'font_color':'white',
                                                'bg_color':'#1F4E78',
                                                'bold': True,
                                                'align':'center',
                                                'valign':'vcenter',
                                                'border':1,
                                                'border_color':'white'
                                                })

            title_format.set_align('vcenter')

            col_format = workbook.add_format({'font_name': 'Arial',
                                                'font_size': 8,
                                                'font_color':'white',
                                                'bg_color':'#1F4E78',
                                                'text_wrap':True,
                                                'border':1,
                                                'border_color':'white',
                                                'align':'center',
                                                'valign':'vcenter'
                                                })

            data_format = workbook.add_format({'font_name': 'Arial',
                                                'font_size': 10,
                                                'align':'left',
                                                'valign':'vcenter'
                                                })
            data_format1 = workbook.add_format({'font_name': 'Arial',
                                                'font_size': 10,
                                                'align':'center',
                                                'valign':'vcenter'
                                                })

            data_format1 = workbook.add_format({'font_name': 'Arial',
                                                'font_size': 10,
                                                'align':'center',
                                                'valign':'vcenter'
                                                })
            num_percent_data_format = workbook.add_format({'font_name':'Arial',
                                                'font_size': 10,
                                                'align':'center',
                                                'valign':'vcenter',
                                                'num_format':'0.00%'
                                                })
            percent_fmt=workbook.add_format({'num_format':'0.00%'})
            statis_format1 = workbook.add_format({'font_name':'Arial',
                                                'font_size': 9,
                                                'align':'left',
                                                'valign':'vcenter',
                                                'bg_color':'#C8EFD0'
                                                })
            statis_format2 = workbook.add_format({'font_name':'Arial',   #系列总计
                                                'font_size': 9,
                                                'align':'center',
                                                'valign':'vcenter',
                                                'bg_color':'#92CDDC'
                                                })
            statis_format4 = workbook.add_format({'font_name': 'Arial',  # 系列总计
                                                  'font_size': 9,
                                                  'align': 'center',
                                                  'valign': 'vcenter',
                                                  'bg_color': '#92CDDC'
                                                  ,'num_format':'0.00%'
                                                  })
            statis_format3 = workbook.add_format({'font_name':'Arial',
                                                'font_size': 9,
                                                'align':'center',
                                                'valign':'vcenter'
                                                })

            item_p=item_report[(item_report['项目号'].str.contains('-...') == False)&(item_report['系列'].str.contains('总计') == False)][['项目号','系列','设备名称','累计下单总项数','实际下单总项数','需求齐料时间','已入库项数','已到货项数','关闭/暂停项数','延期项数','退回返修项数','采购未下单总项数','留置项数']].reset_index(drop=True)
            item_int=['累计下单总项数','实际下单总项数','已入库项数','已到货项数','关闭/暂停项数','延期项数','退回返修项数','采购未下单总项数','留置项数']
            item_p[item_int]=item_p[item_int].fillna(0)
            item_p['累计下单总项数']= pd.to_numeric(item_p['累计下单总项数'],errors='coerce')
            item_p['实际下单总项数']= pd.to_numeric(item_p['实际下单总项数'],errors='coerce')
            item_p['到料量']=item_p['已入库项数']+item_p['已到货项数']+item_p['累计下单总项数']-item_p['实际下单总项数']
            del item_p['已入库项数']
            del item_p['已到货项数']
            del item_p['实际下单总项数']
            report_date=report[(report['使用交期'] != pd.Timestamp(1990, 1, 1))&(report['当前状态'].str.contains("采购"))&(report['使用交期'].dt.date >= datetime.datetime.today().date())]
            #del report['使用交期']
            if len(report_date)>0:
                pivot=pd.pivot_table(report_date,index=[u'项目编号'],columns=[u'使用交期'],values=[u'料件编号'],aggfunc=['count'],fill_value=0,margins=1).reset_index()
                time_col=pd.DataFrame(pd.DataFrame(pivot.columns.tolist())[2])
                time_col.columns=['人生']
                time_col['人生']=time_col['人生'].fillna(default_date)
                time_col['人生'].iloc[-1]=default_date
                time_col['人生']=pd.to_datetime(time_col['人生']).dt.strftime('%Y-%m-%d').astype(str)
                time_col['人生'].iloc[0]='项目编号'
                time_col['人生'].iloc[-1]='总计'
                pivot.columns=time_col['人生'].to_list()
                num_trend=pd.merge(pivot,item_p,left_on='项目编号',right_on='项目号',how='left')
                del num_trend['项目号']
                del num_trend['需求齐料时间']

                #del num_trend['已请购+库存转换总项数']
                t1= num_trend['设备名称']
                t2= num_trend['累计下单总项数']
                #t3= num_trend['实际下单总项数']
                t4=num_trend['退回返修项数']
                t3= num_trend['关闭/暂停项数']
                t5= num_trend['到料量']
                t7= num_trend['延期项数']
                t8= num_trend['采购未下单总项数']
                t9=num_trend['系列']
                del num_trend['设备名称']
                del num_trend['累计下单总项数']
                #del num_trend['实际下单总项数']
                #del num_trend['需求齐料时间']
                del num_trend['到料量']
                del num_trend['退回返修项数']
                del num_trend['关闭/暂停项数']
                del num_trend['延期项数']
                del num_trend['系列']
                del num_trend['采购未下单总项数']
                #num_trend = num_trend.drop(change_col,axis=1)
                num_trend.insert(1,'到料量',t5)
                num_trend.insert(1,'采购未下单总项数',t8)
                num_trend.insert(1,'延期项数',t7)
                num_trend.insert(1,'退回返修项数',t4)
                num_trend.insert(1,'关闭/暂停项数',t3)
                #num_trend.insert(1,'需求齐料时间',t4)
                #num_trend.insert(1,'实际下单总项数',t3)
                num_trend.insert(1,'累计下单总项数',t2)
                num_trend.insert(1,'设备名称',t1)
                num_trend.insert(1,'系列',t9)
                num_trend['到料率']=0
                t6=num_trend['到料率']
                del num_trend['到料率']
                num_trend.insert(9,'到料率',t6)

                num_trend=num_trend.drop(num_trend.tail(1).index)#删除末端总计行
                #每日到料百分比
                per_num_trend=num_trend.copy(deep=True)
                list_col=list(per_num_trend.columns)
                per_num_trend['奋斗']=0
                for i in range(9,(len(list_col)-2)):
                    per_num_trend['奋斗'] = per_num_trend['奋斗'] + per_num_trend[list_col[i]]
                    per_num_trend[list_col[i]]=(per_num_trend['奋斗']+per_num_trend['到料量'])/(per_num_trend['累计下单总项数']- per_num_trend['关闭/暂停项数'])
                   # per_num_trend[list_col[i]]=per_num_trend[list_col[i]].map(lambda x: format(x, '.2%'))
                del per_num_trend['奋斗']
                trend_total=pd.concat([num_trend,per_num_trend]).reset_index(drop=True)
                trend_total=trend_total.sort_values(by=['项目编号','系列','设备名称']).reset_index(drop=True)
                trend_total['到料率']=trend_total['到料量']/(trend_total['累计下单总项数']-trend_total['留置项数']-trend_total['关闭/暂停项数'])
                trend_total.loc[(trend_total['累计下单总项数'] == (trend_total['关闭/暂停项数']+trend_total['留置项数'])),'到料率']=''
                del trend_total['留置项数']
                del item_report['留置项数']
                trend_total_col = ['到料率']

                percent_col_numlist = \
                    [trend_total.columns.tolist().index(i) for i in trend_total_col]
                worksheet4.write_row("A1", trend_total.columns, col_format)
                writer_contents(sheet=worksheet4, array=trend_total.T.values,
                                start_row=1, start_col=0,percent_format=num_percent_data_format,
                                percentlist=percent_col_numlist
                                )
                end = len(trend_total) + 2
                worksheet4.conditional_format('J%s:J%s'%(2,end), {'type': 'data_bar',
                                                        'bar_color': '#4BBB11',
                                                        'data_bar_2010':True,
                                                        'bar_solid':False})
                d=int(len(trend_total))   #
                end=get_column_letter(int((len(list_col)-2)))
                worksheet4.set_column('A:A', 16,data_format)
                worksheet4.set_column('B:B', 14,data_format)
                worksheet4.set_column('C:C', 14,data_format)
                worksheet4.set_column('D:D', 9,data_format1)
                worksheet4.set_column('E:E', 7,data_format1)
                worksheet4.set_column('F:F', 7,data_format1)
                worksheet4.set_column('G:EE', 9,data_format1)
                for i in range(0,d):
                    if i%2==1:
                        m=i+2
                        t=end+str(m)
                        worksheet4.conditional_format('K%s:%s'%(m,t), {'type': 'data_bar',
                                                        'bar_color': '#65C385',
                                                        'data_bar_2010':True,
                                                        'bar_solid':False})
                        worksheet4.conditional_format('K%s:%s' % (m, t), {'type': 'cell',
                                                                          'criteria':'>=',
                                                                          'value':0,'format': percent_fmt})
                lenth=int(len(trend_total))   #表格长度
                col=['项目编号','系列','设备名称','累计下单总项数','关闭/暂停项数','退回返修项数','延期项数','采购未下单总项数','到料量']
                #end=get_column_letter(int((len(list_col)-2)))
                for i in range(1,lenth):
                    for j in range(0,9):
                        if trend_total['项目编号'][i]==trend_total['项目编号'][i-1]:
                            m=get_column_letter(j+1)    #获取列位置位置
                            worksheet4.merge_range('%s:%s'%(m+str(i+1),m+str(i+2)),trend_total[col[j]][i], data_format)
            #形态转换明细
            report002 = report002.reindex(columns=
                    ['项目编号', '库存管理特征',"单据编号","录入日期","资料创建日","申请人员","数据审核日","人员名称","变更类型"
                     ,"料件编号",	"品名","规格","库位","库位说明","变更前-库存管理特征"
                    ,"变更前-库存单位","变更前单位名称","变更后-库存单位","变更后单位名称","变更数量","备注"])
            str002=['项目编号', '库存管理特征',"单据编号","申请人员","人员名称"
                     ,"料件编号",	"品名","规格","库位","库位说明","变更前-库存管理特征"
                       ,"变更前单位名称","变更后单位名称","备注"]
            report002[str002]=report002[str002].fillna('')
            report_xilie=item_report_copy[['项目号','系列','负责人']].drop_duplicates(subset=['项目号']).reset_index(drop=True)
            report_xilie = report_xilie.rename(columns={'项目号': '项目编号'})
            report002=pd.merge(report002,report_xilie,left_on='项目编号',right_on='项目编号',how='left')
            report002=report002.fillna('')
            y1=report002['系列']
            y2=report002['负责人']
            del report002['系列']
            del report002['负责人']
            report002.insert(3,'系列',y1)
            report002.insert(3,'负责人',y2)
            report002=report002.rename(columns={'项目编号':'变更后-项目编号','库存管理特征':'变更后-库存管理特征'})
            date002=["录入日期","资料创建日","数据审核日"]
            report002[date002]=report002[date002].fillna(default_date)
            report002['录入日期'] = pd.to_datetime(report002["录入日期"],errors='coerce').dt.strftime('%Y-%m-%d')
            report002['资料创建日'] = pd.to_datetime(report002["资料创建日"],errors='coerce').dt.strftime('%Y-%m-%d')
            report002['数据审核日'] = pd.to_datetime(report002["数据审核日"],errors='coerce').dt.strftime('%Y-%m-%d')
            report002['录入日期'] = ['' if i == '1990-01-01' else i for i in report002['录入日期']]
            report002['资料创建日'] = ['' if i == '1990-01-01' else i for i in report002['资料创建日']]
            report002['录入日期']  = ['' if i == '1990-01-01' else i for i in report002['录入日期']]

            ###BOM对比
            worksheet3.write_row("A1", report002.columns, col_format)
            writer_contents(sheet=worksheet3, array=report002.T.values,start_row=1, start_col=0)
            worksheet3.set_column('A:A', 15,data_format)
            worksheet3.set_column('B:B', 16,data_format)
            worksheet3.set_column('C:C', 15,data_format)
            worksheet3.set_column('D:D', 9,data_format)
            worksheet3.set_column('E:E', 13,data_format)
            worksheet3.set_column('F:F', 10,data_format)
            worksheet3.set_column('G:G', 10,data_format)
            worksheet3.set_column('H:H', 8,data_format)
            worksheet3.set_column('I:I', 10,data_format)
            worksheet3.set_column('J:J', 8,data_format)
            worksheet3.set_column('K:K', 6,data_format1)
            worksheet3.set_column('L:L', 7,data_format)
            worksheet3.set_column('M:M', 12,data_format)
            worksheet3.set_column('N:N', 10,data_format)
            worksheet3.set_column('O:O', 10,data_format)
            worksheet3.set_column('P:P', 12,data_format)
            worksheet3.set_column('Q:Q', 16,data_format)
            worksheet3.set_column('R:R', 8,data_format1)
            worksheet3.set_column('S:S', 8,data_format)
            worksheet3.set_column('T:T', 8,data_format1)
            worksheet3.set_column('U:U', 8,data_format)
            worksheet3.set_column('V:V', 6,data_format1)
            worksheet3.set_column('W:W', 12,data_format)
            #BOM对比
            ######BOM对比
            report200 = report200.sort_values(by=['项目号', '元件料号']).reset_index(drop=True)
            report200['项目料'] = report200['项目号'].astype(str) + report200['元件料号'].astype(str)
            report200_group = report200.groupby(['项目料'])[['组成用量']].sum().add_suffix('总计').reset_index()
            report200['单设备用料数量'] = pd.merge(report200, report200_group, on='项目料', how='left')['组成用量总计']
            item_report_cop = item_report_cop.drop_duplicates(subset=['项目号']).reset_index()
            report200['设备数量'] = pd.merge(report200, item_report_cop, left_on='项目号', right_on='项目号', how='left')['设备数量']
            report200['BOM数量'] = report200['设备数量'] * report200['单设备用料数量']
            report['项目料'] = report['项目编号'].astype(str) + report['料件编号'].astype(str)
            repor500 = report[report['当前状态'].str.contains('未下单') == False].reset_index(drop=True)
            repor929 = report[report['当前状态'].str.contains('未下单')].reset_index(drop=True)
            report_group500 = repor500.groupby(['项目料'])[['需求数量']].sum().add_suffix('总计').reset_index()
            report_group929 = repor929.groupby(['项目料'])[['需求数量']].sum().add_suffix('总计').reset_index()
            if len(report_group500)==0:
                report_group500=pd.DataFrame(columns=['项目料','需求数量总计'])
            if len(report_group929)==0:
                report_group929=pd.DataFrame(columns=['项目料','需求数量总计'])
            print(report_group500)
            print(report_group929)
            report200['请购总数量'] = 0
            report200['采购已下单数量'] = pd.merge(report200, report_group500, on='项目料', how='left')['需求数量总计']
            report200['采购未下单数量'] = pd.merge(report200, report_group929, on='项目料', how='left')['需求数量总计']
            report002['项目料']=report002['变更后-项目编号'].astype(str) + report002['料件编号'].astype(str)
            if len(report002)>0:
                report002_group=report002.groupby(['项目料'])[['变更数量']].sum().add_suffix('总计').reset_index()
            else:
                report002_group=pd.DataFrame(columns=['项目料','变更数量总计'])
            report200['形态转换数量']=pd.merge(report200,report002_group,how='left',on='项目料')['变更数量总计']
            cpmq012_use['项目料']=cpmq012_use['项目号'].astype(str)+cpmq012_use['料号'].astype(str)
            cpmq012_use['项目料']=cpmq012_use['项目料'].fillna('')
            if len(cpmq012_use)>0:
                cpmq012_use_group = cpmq012_use.groupby(['项目料'])[['在途占用未入库数量']].sum().add_suffix('').reset_index()
            else:
                cpmq012_use_group =pd.DataFrame(columns=['项目料','在途占用未入库数量'])
            report200['在途占用公共仓数量'] = pd.merge(report200, cpmq012_use_group, how='left', on='项目料')['在途占用未入库数量']
            for num in ['采购已下单数量','采购未下单数量','形态转换数量','在途占用公共仓数量']:
                report200[num]=report200[num].fillna(0)
            report200['请购总数量']=report200['采购已下单数量']+report200['采购未下单数量']+report200['形态转换数量']+report200['在途占用公共仓数量']
            report200['差异']=report200['BOM数量']-report200['请购总数量']
            report200['备注']=''
            report200=report200.fillna('')

            del cpmq012_use['项目料']
            del report['项目料']
            del report200['项目料']
            worksheet1_3.write_row("A1", report200.columns, col_format)
            writer_contents(sheet=worksheet1_3, array=report200.T.values, start_row=1, start_col=0)
            # 物控表填充数据并设置格式
            del report['请购天数']

            report=report.reset_index(drop=True)
            report['请购日期'] = pd.to_datetime(report["请购日期"], errors='coerce').dt.strftime('%Y-%m-%d')
            report['到货日期'] = pd.to_datetime(report["到货日期"], errors='coerce').dt.strftime('%Y-%m-%d')
            report['采购日期'] = pd.to_datetime(report["采购日期"], errors='coerce').dt.strftime('%Y-%m-%d')
            report['采购确认日期'] = pd.to_datetime(report["采购确认日期"], errors='coerce').dt.strftime('%Y-%m-%d')
            report['需求日期'] = pd.to_datetime(report["需求日期"], errors='coerce').dt.strftime('%Y-%m-%d')
            report['供应商回复交期'] = pd.to_datetime(report["供应商回复交期"], errors='coerce').dt.strftime('%Y-%m-%d')
            report['使用交期'] = pd.to_datetime(report["使用交期"], errors='coerce').dt.strftime('%Y-%m-%d')
            report['数据审核日'] = pd.to_datetime(report["数据审核日"], errors='coerce').dt.strftime('%Y-%m-%d')
            report['最新需求日期'] = pd.to_datetime(report["最新需求日期"], errors='coerce').dt.strftime('%Y-%m-%d') #'已检验数量','入库日期'
            report['入库日期'] = pd.to_datetime(report["入库日期"], errors='coerce').dt.strftime('%Y-%m-%d')
            report[['请购日期','采购日期','采购确认日期','供应商回复交期','数据审核日','需求日期','最新需求日期','入库日期',"使用交期"]] =\
                report[['请购日期','采购日期','采购确认日期','供应商回复交期','数据审核日','需求日期','最新需求日期','入库日期',"使用交期"]].astype(str)
            report['请购日期'] = ['' if i == '1990-01-01' else i
                              for i in report['请购日期']]
            report['采购日期'] = ['' if i == '1990-01-01' else i
                              for i in report['采购日期']]
            report['采购确认日期'] = ['' if i == '1990-01-01' else i
                              for i in report['采购确认日期']]
            report['供应商回复交期'] = ['' if i == '1990-01-01' else i
                              for i in report['供应商回复交期']]
            report['数据审核日'] = ['' if i == '1990-01-01' else i
                                 for i in report['数据审核日']]
            report['需求日期'] = ['' if i == '1990-01-01' else i
                              for i in report['需求日期']]
            report['最新需求日期'] = ['' if i == '1990-01-01' else i
                              for i in report['最新需求日期']]
            report['到货日期'] = ['' if i == '1990-01-01' else i
                                for i in report['到货日期']]
            report['入库日期'] = ['' if i == '1990-01-01' else i for i in report['入库日期']]
            report['使用交期'] = ['' if i == '1990-01-01' else i for i in report['使用交期']]
            #report['CEG确认日期'] = ['' if i == '1990-01-01' else i
             #                 for i in report['CEG确认日期']]
            report['交期差异天数'] = report['交期差异天数'].fillna('')
            report['设计延期天数'] = report['设计延期天数'].fillna('')

            report.loc[(report['采购确认日期']=='2088-01-01')&(report['当前状态'].str.contains('采购中')),'当前状态'] = '通知送货'
            report['作业编号'] = report['作业编号'].fillna('')

            report.loc[(report['作业编号']) == '无模组号', '作业编号'] = ''

            report=report.fillna('')
            report=report.rename(columns={'使用交期':'系统交期汇总'})
            report = report.reindex(columns=
                                    ['项目编号', '库存管理特征', '作业编号', '模组名称', 'P工序', '系列', '设备',
                                     '项目阶段',
                                     '负责人', '请购日期',  '来源单号', '来源项次',
                                     '物料请购负责人', '状态码', '行状态', '料件编号',
                                     '品名', '规格', '需求数量', '品牌', '需求日期', '最新需求日期', '零件类型',
                                     '表面处理',
                                     '采购负责人',
                                     '采购日期', '数据审核日', '采购单号', '采购项次', '供应商名称', '采购确认日期',
                                     '供应商回复交期',
                                     '已收货量', '已入库量', '仓退数量', '验退数量', '未交量',
                                     '当前状态', '风险等级', '类别分类', '采购组别', '唯一值', '交期差异天数',
                                     '超3天未下单', '交期判断', '已检验数量', '入库日期', '到货日期','系统交期汇总',
                                     '物流信息', '大项目名称',
                                     '采购延期说明', '延期进度说明', '是否CEG确认', 'CEG备注', 'CEG确认货期(天)',
                                     '设计备注', '交货地址栏(请购单)', '交货地址栏(采购单)', '请购单备注',
                                     '采购确认日期履历', '最新需求日期变更履历', 'MC负责人', '核价采购员姓名',
                                     '退货快递单号', 'MC风险判断', '录入需求日期', '是否大件物料', '设计延期天数',
                                     '理由码说明','旧版料号','备注说明'])

            worksheet1.write_row("A1", report.columns, col_format)
            # report.loc[:, '物料需求交期(模组装配计划减生产准备前置期)']=convert_exceldate_col(report['物料需求交期(模组装配计划减生产准备前置期)'])
            writer_contents(sheet=worksheet1, array=report.T.values, start_row=1,start_col=0)
            write_color(book=workbook, sheet=worksheet1, data=report['当前状态'],fmt=data_format1, col_num='AL')
            write_color1(book=workbook, sheet=worksheet1, data=report['风险等级'],fmt=data_format1, col_num='AM')

            #worksheet1format(worksheet1)
                # 整体数据汇总填充数据
            if '留置项数' in item_report.columns:
                del item_report['留置项数']
            worksheet2.merge_range('A1:AX1', 'XXXXX项目物料整体数据一览表', title_format)
            worksheet2.merge_range('A2:G3', '项目情况', col_format)
            worksheet2.merge_range('H2:AD2', '设计下单情况', col_format)
            worksheet2.merge_range('AE2:AF3', 'MC抛单情况', col_format)
            worksheet2.merge_range('AG2:AL2', '采购下单情况', col_format)
            worksheet2.merge_range('AM2:AX3', '回料状态', col_format)
            worksheet2.merge_range('H3:M3', '1参考BOM', col_format)
            worksheet2.merge_range('N3:S3', '2设计计划', col_format)
            worksheet2.merge_range('T3:X3', '3 T100 本机BOM', col_format)
            worksheet2.merge_range('Z3:AD3','实际系统已请未购+采购明细+形态转换', col_format)
            worksheet2.merge_range('AG3:AI3', '采购已下单', col_format)
            worksheet2.merge_range('AJ3:AL3', '采购未下单', col_format)
            worksheet2.write_row('A4', item_report.columns, col_format)
            worksheet2.write('Y3', 'M3抛转数据', col_format)

            #item_report["计划出货日期"] = pd.to_datetime(item_report["计划出货日期"]).dt.strftime('%Y-%m-%d')
            item_report['需求齐料时间'] = pd.to_datetime(item_report["需求齐料时间"],errors='coerce').dt.strftime('%Y-%m-%d')
            item_report['计划出货日期'] = pd.to_datetime(item_report["计划出货日期"],errors='coerce').dt.strftime('%Y-%m-%d')
            item_report['设计下单完成日期'] = pd.to_datetime(item_report["设计下单完成日期"],errors='coerce').dt.strftime('%Y-%m-%d')
            item_report['到料95%交期'] = pd.to_datetime(item_report['到料95%交期'],errors='coerce').dt.strftime('%Y-%m-%d')
            item_report_date_col = ['需求齐料时间', '计划出货日期', '设计下单完成日期','到料95%交期']
            default_date = pd.Timestamp(1990, 1, 1)
            item_report[item_report_date_col] = \
                item_report[item_report_date_col].fillna(default_date)
            item_report['需求齐料时间'] = pd.to_datetime(item_report["需求齐料时间"],errors='coerce').dt.strftime('%Y-%m-%d')
            item_report['计划出货日期'] = pd.to_datetime(item_report["计划出货日期"],errors='coerce').dt.strftime('%Y-%m-%d')
            item_report['设计下单完成日期'] = pd.to_datetime(item_report["设计下单完成日期"],errors='coerce').dt.strftime('%Y-%m-%d')
            item_report['到料95%交期'] = pd.to_datetime(item_report['到料95%交期'],errors='coerce').dt.strftime('%Y-%m-%d')
            item_report[['需求齐料时间','计划出货日期','设计下单完成日期','到料95%交期']] =\
                item_report[['需求齐料时间','计划出货日期','设计下单完成日期','到料95%交期']].astype(str)
            item_report['需求齐料时间']= ['' if i == '1990-01-01' else i
                              for i in item_report['需求齐料时间']]
            item_report['计划出货日期']= ['' if i == '1990-01-01' else i
                              for i in item_report['计划出货日期']]
            item_report['设计下单完成日期']= ['' if i == '1990-01-01' else i
                              for i in item_report['设计下单完成日期']]
            item_report['到料95%交期']= ['' if i == '1990-01-01' else i
                              for i in item_report['到料95%交期']]
            item_report_percent_col = ['超计划日期下单比例', '实际下单进度', '抛单进度%','采购下单率',
                                           '超3天未下单比例'   #, '延期比例'
                                 ,'到料率', '入库率','实际到料率（考虑设计下单进度）','实际入库率（考虑设计下单进度）']
            percent_col_numlist = \
                [item_report.columns.tolist().index(i) for i in item_report_percent_col]

            writer_contents(sheet=worksheet2, array=item_report.T.values,
                            start_row=4, start_col=0,
                            percent_format=num_percent_data_format,
                            percentlist=percent_col_numlist)

            row_count=len(item_report)

            for row_index in range(row_count):
                if '总计' in str(item_report.iloc[row_index,2]):
                    worksheet2.write_row(row_index+4,0,item_report.iloc[row_index,:27].reset_index(drop=True).values,statis_format2)

                    worksheet2.write_row(row_index + 4, 27, item_report.iloc[row_index,27:29].reset_index(drop=True).values,
                                         statis_format4)
                    worksheet2.write_row(row_index + 4, 29, item_report.iloc[row_index, 29:31].reset_index(drop=True).values,
                                         statis_format2)
                    worksheet2.write_row(row_index + 4, 31,
                                         item_report.iloc[row_index, 31:32].reset_index(drop=True).values,
                                         statis_format4)
                    worksheet2.write_row(row_index + 4, 32, item_report.iloc[row_index, 32:33].reset_index(drop=True).values,
                                         statis_format2)
                    worksheet2.write_row(row_index + 4, 33,
                                         item_report.iloc[row_index, 33:34].reset_index(drop=True).values,
                                         statis_format4)
                    worksheet2.write_row(row_index + 4, 34,
                                         item_report.iloc[row_index, 34:37].reset_index(drop=True).values,
                                         statis_format2)
                    worksheet2.write_row(row_index + 4, 37,
                                         item_report.iloc[row_index, 37:38].reset_index(drop=True).values,
                                         statis_format4)
                    worksheet2.write_row(row_index + 4, 38,
                                         item_report.iloc[row_index, 38:45].reset_index(drop=True).values,
                                         statis_format2)
                    worksheet2.write_row(row_index + 4, 45,
                                         item_report.iloc[row_index, 45:47].reset_index(drop=True).values,
                                         statis_format4)
                    worksheet2.write_row(row_index + 4, 47,
                                         item_report.iloc[row_index, 47:48].reset_index(drop=True).values,
                                         statis_format2)
                    worksheet2.write_row(row_index + 4, 48,
                                         item_report.iloc[row_index, 48:50].reset_index(drop=True).values,
                                         statis_format4)

            # 列条件格式
            start = 4
            end = len(item_report) + start+250
            worksheet2.conditional_format('AB%s:AB%s'%(start,end), {'type': 'data_bar',
                                                    'bar_color': '#D6007B',
                                                    'data_bar_2010':True,
                                                    'bar_solid':False})
            worksheet2.conditional_format('AC%s:AC%s'%(start,end), {'type': 'data_bar',
                                                    'bar_color': '#FF5B60',
                                                    'data_bar_2010':True,
                                                    'bar_solid':False})
            worksheet2.conditional_format('AF%s:AF%s'%(start,end), {'type': 'data_bar',
                                                    'bar_color': '#FDFE00',
                                                    'data_bar_2010':True,
                                                    'bar_solid':False})
            worksheet2.conditional_format('AH%s:AH%s' % (start, end), {'type': 'data_bar',
                                                                       'bar_color': '#3399FF',
                                                                       'data_bar_2010': True,
                                                                       'bar_solid': False})
            worksheet2.conditional_format('AL%s:AL%s'%(start,end), {'type': 'data_bar',
                                                    'bar_color': '#FFC000',
                                                    'data_bar_2010':True,
                                                    'bar_solid':False})
            worksheet2.conditional_format('AT%s:AT%s'%(start,end), {'type': 'data_bar',
                                                    'bar_color': '#C4D79B',
                                                    'data_bar_2010':True,
                                                    'bar_solid':False})
            worksheet2.conditional_format('AU%s:AU%s' % (start, end), {'type': 'data_bar',
                                                                       'bar_color': '#00B050',
                                                                       'data_bar_2010': True,
                                                                       'bar_solid': False})
            worksheet2.conditional_format('AW%s:AW%s' % (start, end), {'type': 'data_bar',
                                                                       'bar_color': '#B1A0C7',
                                                                       'data_bar_2010': True,
                                                                       'bar_solid': False})
            worksheet2.conditional_format('AX%s:AX%s' % (start, end), {'type': 'data_bar',
                                                                       'bar_color': '#FF9999',
                                                                       'data_bar_2010': True,
                                                                       'bar_solid': False})

            worksheet2.set_column('A:A', 7,data_format1)
            worksheet2.set_column('B:B', 22,data_format)
            worksheet2.set_column('C:C', 16,data_format)
            worksheet2.set_column('D:D', 12,data_format)
            worksheet2.set_column('E:E', 10,data_format)
            worksheet2.set_column('F:F', 7,data_format1)
            worksheet2.set_column('G:G', 10,data_format)
            worksheet2.set_column('H:H', 10,data_format, {'hidden': True})
            worksheet2.set_column('I:I', 8,data_format1, {'hidden': True})
            worksheet2.set_column('J:J', 8,data_format1, {'hidden': True})
            worksheet2.set_column('K:K', 8,data_format1, {'hidden': True})
            worksheet2.set_column('L:L', 8,data_format1, {'hidden': True})
            worksheet2.set_column('M:M', 8,data_format1, {'hidden': True})
            worksheet2.set_column('N:N', 8,data_format1)
            worksheet2.set_column('O:O', 8,data_format1, {'hidden': True})
            worksheet2.set_column('P:P', 8,data_format1, {'hidden': True})
            worksheet2.set_column('Q:Q', 8,data_format1, {'hidden': True})
            worksheet2.set_column('R:R', 8,data_format1, {'hidden': True})
            worksheet2.set_column('S:S', 10,data_format)
            worksheet2.set_column('T:T', 10,data_format1)
            worksheet2.set_column('U:U', 8,data_format1, {'hidden': True})
            worksheet2.set_column('V:V', 8,data_format1, {'hidden': True})
            worksheet2.set_column('W:W', 8,data_format1, {'hidden': True})
            worksheet2.set_column('X:X', 8,data_format1, {'hidden': True})
            worksheet2.set_column('Y:Y', 8,data_format1)
            worksheet2.set_column('Z:Z', 8,data_format1)
            worksheet2.set_column('AA:AA', 8,data_format1)
            worksheet2.set_column('AB:AB', 8,num_percent_data_format)
            worksheet2.set_column('AC:AC', 9,num_percent_data_format)
            worksheet2.set_column('AD:AD', 8,data_format1)
            worksheet2.set_column('AE:AE', 9,data_format1)
            worksheet2.set_column('AF:AF', 8,num_percent_data_format)
            worksheet2.set_column('AG:AG', 8,data_format1)
            worksheet2.set_column('AH:AH', 8,num_percent_data_format)
            worksheet2.set_column('AI:AI', 9,data_format1)
            worksheet2.set_column('AJ:AJ', 8,data_format1)
            worksheet2.set_column('AK:AK', 8,data_format1)
            worksheet2.set_column('AL:AL', 8,num_percent_data_format)
            worksheet2.set_column('AM:AM', 7,data_format1)
            worksheet2.set_column('AN:AN', 7,data_format1)
            worksheet2.set_column('AO:AO', 8,data_format1)
            worksheet2.set_column('AP:AP', 9,data_format1)
            worksheet2.set_column('AQ:AQ', 9,data_format1)
            worksheet2.set_column('AR:AR', 10,data_format1)
            worksheet2.set_column('AS:AS', 10,data_format1)
            worksheet2.set_column('AT:AT', 10,num_percent_data_format)
            worksheet2.set_column('AU:AU', 10,num_percent_data_format)
            worksheet2.set_column('AW:AW', 10,num_percent_data_format)
            worksheet2.set_column('AX:AX', 10,num_percent_data_format)

            worksheet5.write_row("A1", cpmq012_use.columns, col_format)
            writer_contents(sheet=worksheet5, array=cpmq012_use.T.values, start_row=1, start_col=0)


            workbook.close()
            print(bookname+'已完成{:.0%}'.format(1.0))

            screm.insert(INSERT, '\n'+bookname+'已完成{:.0%}'.format(1.0))
            window.update()
############################################################################################################################################################################################################
        #需求项目列表
            #######cpmq012
        screm.insert(INSERT, '\n正在处理在途占用公共仓明细表...', '\n')
        window.update()
        path = r'数据源'
        path_file = os.listdir(path)
        for file in path_file:
            if 'cpmq012' in file and 'xlsx' not in file:
                cpmq012_path = r'数据源\cpmq012'
                cpmq012_file = os.listdir(cpmq012_path)
                for i in cpmq012_file:
                    if '~$' in i:
                        cpmq012_file.remove(i)
                if os.listdir(cpmq012_path):
                    for i in range(len(cpmq012_file)):
                        if str(cpmq012_file[i]).count('~$') == 0:
                            cpmq012 = pd.read_excel(cpmq012_path + '\\' + cpmq012_file[i])
                    cpmq012 = cpmq012[["请购单号", "请购项次", "请购日期", "物料请购负责人", "核价采购员", "核价采购员名称",
                                       "交货地址栏(请购单)", "料号", "品名"
                        , "规格", "项目号", "请购数量", "在途占用数量", "公共项目请购单号", "公共项目请购项次",
                                       "公共项目请购数量", "公共请购单状态"
                        , "在途占用已入库数量", "公共项目号", "公共项目采购单号", "采购项次", "采购日期", "采购人员",
                                       "采购人员名称", "开单人", "开单人名称"
                        , "交货日期", "品牌", "采购确认日期", "供应商回复交期", "供应商", "供应商名称", "采购延期说明",
                                       "物流信息", "到货日期", "交货地址栏(采购单)"
                        , "CEG备注(采购单)", "公共项目采购数量", "已收货数量", "公共采购已入库数量", "AMRP运行时间",
                                       "转单人", "转单人名称", "转单时间"]]
                    text_012 = ["请购单号", "请购项次", "物料请购负责人", "核价采购员", "核价采购员名称",
                                "交货地址栏(请购单)", "料号", "品名"
                        , "规格", "项目号", "公共项目请购单号", "公共请购单状态", "公共项目号", "公共项目采购单号",
                                "采购人员", "采购人员名称"
                        , "开单人", "开单人名称", "品牌", "供应商", "供应商名称", "采购延期说明", "物流信息", "交货地址栏(采购单)"
                        , "CEG备注(采购单)", "转单人", "转单人名称"]
                    num_012 = ["请购项次", "请购数量", "在途占用数量", "公共项目请购项次", "公共项目请购数量"
                        , "在途占用已入库数量", "采购项次", "公共项目采购数量", "已收货数量", "公共采购已入库数量"]
                    date_012 = ["请购日期", "采购日期", "交货日期", "采购确认日期", "供应商回复交期", "到货日期",
                                "AMRP运行时间", "转单时间"]
                    cpmq012[text_012] = cpmq012[text_012].fillna('')
                    cpmq012 = cpmq012[cpmq012['公共请购单状态'].str.contains('结案') == False].reset_index(drop=True)
                    cpmq012[num_012] = cpmq012[num_012].fillna(0)
                    default_date=pd.Timestamp(1990, 1, 1)
                    cpmq012[date_012] = cpmq012[date_012].fillna(default_date)
                    for dat in date_012:
                        cpmq012[dat] = pd.to_datetime(cpmq012[dat], errors='coerce').dt.strftime('%Y-%m-%d').astype(str)
                        cpmq012[dat] = ['' if i == '1990-01-01' else i for i in cpmq012[dat]]
                else:
                    cpmq012=pd.DataFrame(columns=["请购单号", "请购项次", "请购日期", "物料请购负责人", "核价采购员", "核价采购员名称",
                                       "交货地址栏(请购单)", "料号", "品名"
                        , "规格", "项目号", "请购数量", "在途占用数量", "公共项目请购单号", "公共项目请购项次",
                                       "公共项目请购数量", "公共请购单状态"
                        , "在途占用已入库数量", "公共项目号", "公共项目采购单号", "采购项次", "采购日期", "采购人员",
                                       "采购人员名称", "开单人", "开单人名称"
                        , "交货日期", "品牌", "采购确认日期", "供应商回复交期", "供应商", "供应商名称", "采购延期说明",
                                       "物流信息", "到货日期", "交货地址栏(采购单)"
                        , "CEG备注(采购单)", "公共项目采购数量", "已收货数量", "公共采购已入库数量", "AMRP运行时间",
                                       "转单人", "转单人名称", "转单时间"])
        if 'cpmq012' not in path_file:
            cpmq012 = pd.DataFrame(
                columns=["请购单号", "请购项次", "请购日期", "物料请购负责人", "核价采购员", "核价采购员名称",
                         "交货地址栏(请购单)", "料号", "品名"
                    , "规格", "项目号", "请购数量", "在途占用数量", "公共项目请购单号", "公共项目请购项次",
                         "公共项目请购数量", "公共请购单状态"
                    , "在途占用已入库数量", "公共项目号", "公共项目采购单号", "采购项次", "采购日期",
                         "采购人员",
                         "采购人员名称", "开单人", "开单人名称"
                    , "交货日期", "品牌", "采购确认日期", "供应商回复交期", "供应商", "供应商名称",
                         "采购延期说明",
                         "物流信息", "到货日期", "交货地址栏(采购单)"
                    , "CEG备注(采购单)", "公共项目采购数量", "已收货数量", "公共采购已入库数量", "AMRP运行时间",
                         "转单人", "转单人名称", "转单时间"])
        ############判断是否有补充日期表
        screm.insert(INSERT, '\n正在处理委外批次作业表...', '\n')
        window.update()
        path = r'数据源'
        path_file = os.listdir(path)
        for file in path_file:
            if 'asfp400' in file and 'xlsx' not in file:
                asfp400_path = r'数据源\asfp400'
                # need_price = ['料件编号', '最新价格']
                asfp400_file = os.listdir(asfp400_path)
                for i in asfp400_file:
                    if '~$' in i:
                        asfp400_file.remove(i)
                asfp400_1 = []
                if os.listdir(asfp400_path):
                    for i in range(len(asfp400_file)):
                        if str(asfp400_file[i]).count('~$') == 0:
                            asfp400_2 = pd.read_excel(asfp400_path + '\\' + asfp400_file[i])  # 补充需求日期
                            asfp400_1.append(asfp400_2)
                    asfp400 = pd.concat(asfp400_1).reset_index(drop=True)
                    need_400 = ['选择', '单号', '核价制单人名称','工单日期', '核价采购员名称', '生产料号', '项目编号',
                                '项目名称', '品名', '规格', '可委外数量', '备注', '交货地址', '设计备注',
                                '大项目名称', '工单发料料号']

                    for col in asfp400.columns:
                        if col not in need_400:
                            del asfp400[col]

                    asfp400 = asfp400[need_400].fillna('')
                    asfp400['工单日期'] = asfp400['工单日期'].fillna(default_date)
                    asfp400['工单日期'] = pd.to_datetime(asfp400['工单日期'], errors='coerce').dt.strftime(
                        '%Y-%m-%d').astype(str)
                    asfp400['工单日期'] = ['' if i == '1990-01-01' else i
                                           for i in asfp400['工单日期']]
                    asfp400.loc[asfp400['核价制单人名称'] == '', '核价制单人名称'] = asfp400['核价采购员名称']
                    del asfp400['核价采购员名称']
                    asfp400 = asfp400.rename(
                        columns={'选择': '是否CEG确认', '单号': '来源单号', '核价制单人名称': '核价采购员姓名',
                                 '生产料号': '料件编号', '项目名称': '设备', '可委外数量': '需求数量',
                                 '备注': '请购单备注', '交货地址': '交货地址栏(请购单)',
                                 '工单发料料号': '旧版料号','工单日期':'请购日期'})
                    asfp400['当前状态']='未下单'
                    asfp400 = asfp400.fillna('')
                    asfp400=asfp400[['是否CEG确认','来源单号',  '核价采购员姓名','料件编号', '项目编号', '设备','品名','规格', '需求数量', '请购单备注','交货地址栏(请购单)','设计备注','大项目名称','旧版料号','当前状态','请购日期']]
                else:
                    asfp400 =pd.DataFrame(columns=['是否CEG确认','来源单号','请购日期',  '核价采购负责人','料件编号', '项目编号', '设备','品名','规格', '需求数量', '请购单备注','交货地址栏(请购单)','设计备注','大项目名称','旧版料号','当前状态','请购日期'])
        if 'asfp400' not in path_file:
            asfp400 = pd.DataFrame(columns=['是否CEG确认', '来源单号', '核价采购员姓名', '料件编号','请购日期', '项目编号','设备', '品名', '规格', '需求数量',
                         '请购单备注', '交货地址栏(请购单)', '设计备注', '大项目名称', '旧版料号', '当前状态'])
        item_report_all= pd.ExcelFile('数据源/需求项目列表.xlsx')
        for name in item_report_all.sheet_names:
            item_report=item_report_all.parse(name)
            item_report['项目号']=item_report['项目号'].fillna('空')
            item_report_copy=item_report.copy()
            item_report_copy=item_report_copy.fillna('')
            item_report_copy = item_report_copy.rename(columns={'项目负责人': '负责人'})
            item_report=item_report[item_report['项目号']!='空'].reset_index(drop=True)
            add_report(item_report, name)
    except Exception  as e:
        print('异常信息为:', e)  # 异常信息为: division by zero
        screm.insert(INSERT, '\n---------------------程序报错，异常信息为:'+traceback.format_exc())

    end_time = time.time()

    print('程序总耗时:%d秒' % (end_time-start_time))
    print('已完成{:.0%}'.format(1))
    screm.insert(INSERT, '\n程序总耗时:%d秒' % (end_time-start_time))
    screm.insert(INSERT, '\n全部已完成{:.0%}'.format(1))
    window.update()
#a2.bind("<<ComboboxSelected>>",func)  #等同于textvariable=cv这个变量
button_execute['command']=execute

#加工工具按钮
add_deal= tk.Button(frame,
              text="加工操作",width=8, height=1, fg='navy',bd=4,font=('华文行楷', 12))              #command=createNewWindow
frame.create_window(50,110,window=add_deal)


def createNewWindow():
    def upload_file1():
        selectFile = tk.filedialog.askopenfilename()  # askopenfilename 1次上传1个；askopenfilenames1次上传多个
        entry1.insert(0, selectFile)

    def upload_file2():
        selectFile = tk.filedialog.askopenfilename()  # askopenfilename 1次上传1个；askopenfilenames1次上传多个
        entry2.insert(0, selectFile)

    def upload_file3():
        selectFile = tk.filedialog.askopenfilename()  # askopenfilename 1次上传1个；askopenfilenames1次上传多个
        entry3.insert(0, selectFile)
    newWindow = tk.Toplevel(window)
    newWindow.geometry('600x600')
    newWindow.title('物控表加工操作')
    frm = tk.Frame(newWindow)
    frm.grid(padx='20', pady='30')

    btn1 = tk.Button(frm, text='上传文件（当前物控表）', command=upload_file1)
    btn1.grid(row=0, column=0, ipadx='3', ipady='3', padx='10', pady='20')
    entry1 = tk.Entry(frm, width=50, textvariable=tk.StringVar())
    entry1.grid(row=0, column=1)
    btn2 = tk.Button(frm, text='上传文件（cpjt200）', command=upload_file2)
    btn2.grid(row=3, column=0, ipadx='3', ipady='3', padx='10', pady='20')
    entry2 = tk.Entry(frm, width=50, textvariable=tk.StringVar())
    entry2.grid(row=3, column=1)
    btn3 = tk.Button(frm, text='上传文件（之前物控表）', command=upload_file3)
    btn3.grid(row=6, column=0, ipadx='3', ipady='3', padx='10', pady='20')
    entry3 = tk.Entry(frm, width=50, textvariable=tk.StringVar())
    entry3.grid(row=6, column=1)
    btn4 = tk.Button(frm, text='执 行', bg="green", fg="white")
    btn4.grid(row=7, column=1, ipadx='1', ipady='1', padx='8', pady='3')

    def countifs(criteria_ranges, criterias):
        try:
            count = criteria_ranges.value_counts().loc[criterias]
        except:
            count = 0
        return count

    def excel():
        errorscrem = tk.Text(frm, bg='white',  # 标签背景颜色
                             font=('微软雅黑', 9),  # 字体和字体大小
                             width=56, height=16,  # 标签长宽(以字符长度计算)
                             )
        errorscrem.grid(row=8, column=1,sticky=E)
        # 之前物控表
        try:
            workbook_oldall = pd.read_excel(entry3.get(), sheet_name=None)
            workbook_old = workbook_oldall['物料管控表']
            if '项次' in workbook_old:
                workbook_old = workbook_old.rename(columns={'项次': '采购项次'})
            workbook_old['采购项次'] = workbook_old['采购项次'].fillna(0)
            workbook_old['采购项次'] = pd.to_numeric(workbook_old['采购项次'])
            workbook_old['采购项次'] = workbook_old['采购项次'].fillna(0).round(0).astype(int)
            workbook_old['超链码'] = workbook_old['项目编号'].astype(str) + workbook_old['采购项次'].astype(str) + workbook_old[
                '采购单号'].astype(str) + workbook_old['作业编号'].astype(str) + workbook_old['料件编号'].astype(str) + workbook_old['来源单号'].astype(str)
            workbook_old=workbook_old.drop_duplicates(subset=['超链码']).reset_index(drop=True)
            # 今天物控表
            workbook_nowall = pd.read_excel(entry1.get(), sheet_name=None)
            workbook_report=workbook_nowall['物料管控表']
            if len(workbook_report)>0:
                workbook_report['采购项次'] = pd.to_numeric(workbook_report['采购项次'])
                workbook_report['采购项次'] = workbook_report['采购项次'].fillna(0).round(0).astype(int)
                workbook_report['超链码'] = workbook_report['项目编号'].astype(str) + workbook_report['采购项次'].astype(str) + \
                                     workbook_report['采购单号'].astype(str) + workbook_report['作业编号'].astype(str) + \
                                     workbook_report['料件编号'].astype(str) + workbook_report['来源单号'].astype(str)
                workbook_item = pd.read_excel(entry1.get(), header=3, sheet_name='整体数据汇总')
                now_time = time.strftime("%Y-%m-%d-%H-%M", time.localtime(time.time()))
                book_name = '加工物控表' + now_time
                shutil.copy(entry1.get(), os.path.join(os.getcwd(), book_name + '.xlsx'))
                '''
                default_date = '1990/01/01'
                workbook_report['最新需求日期-200'] = default_date
                workbook_report['最新需求日期-200'] = pd.to_datetime(workbook_report['最新需求日期-200'], errors='coerce')
                # workbook_report['最新需求日期-200'] =workbook_report['最新需求日期-200'].fillna(default_date)
                '''
                workbook200 = pd.read_excel(entry2.get())  # cpjt200
                project_list = list(workbook_item['项目号'])
                workbook200 = workbook200[
                    sum(workbook200['项目编号'].str.contains(project) for project in project_list) > 0].reset_index(drop=True)
                workbook200.loc[workbook200['作业编号'] == " ", '作业编号'] = '无模组号'
                workbook200['作业编号'] = workbook200['作业编号'].fillna('无模组号')
                workbook200['库存管理特征'] = workbook200['项目编号'] + '-' + workbook200['作业编号']
                '''
                workbook200['需求期1'] = ''
                int_col = '生产准备前置期/天'
                workbook200[int_col] = workbook200[int_col].fillna(0)
                workbook200.loc[workbook200['生产准备前置期/天'] == " ", '生产准备前置期/天'] = 0
                workbook200['生产准备前置期/天'] = pd.to_numeric(workbook200['生产准备前置期/天'])
                workbook200['模组装配日期'] = workbook200['模组装配日期'].fillna(default_date)
                workbook200.loc[workbook200['模组装配日期'] == '          ', '模组装配日期'] = default_date
                workbook200['模组装配日期'] = pd.to_datetime(workbook200["模组装配日期"], errors='coerce')
                workbook200 = workbook200.reset_index(drop=True)
                workbook200['生产准备前置期/天'] = pd.to_numeric(workbook200['生产准备前置期/天'])
                workbook200['需求期1'] = workbook200['需求期1'].fillna(default_date)
                workbook200['需求期1'] = pd.to_datetime(workbook200['模组装配日期']) - pd.to_timedelta(workbook200['生产准备前置期/天'],
                                                                                              unit='D')
                workbook200['需求期1'] = workbook200['需求期1'].fillna(default_date)
                workbook_report = workbook_report.reset_index(drop=True)
                workbook4 = workbook200.drop_duplicates(subset=['库存管理特征'], keep='last').reset_index(drop=True)
                workbook_report['最新需求日期-200'] = \
                    pd.merge(workbook_report, workbook4, left_on='库存管理特征', right_on='库存管理特征', how='left').reset_index(
                        drop=True)[
                        '需求期1']
                workbook_report['最新需求日期-200'] = workbook_report['最新需求日期-200'].fillna(default_date)
                workbook_report['最新需求日期-200'] = pd.to_datetime(workbook_report["最新需求日期-200"], errors='coerce')
                '''
                # 模组名称
                workbook200 = workbook200.rename(columns={'模组名称': '模组名称200'})
                workbook200['模组名称200'] = workbook200['模组名称200'].fillna('abc')
                workbook3 = workbook200[workbook200['模组名称200'] != 'abc'][['库存管理特征', '模组名称200']].drop_duplicates(subset=['库存管理特征'],keep='first').reset_index(drop=True)
                workbook_report = workbook_report.reset_index(drop=True)
                workbook3 = workbook3.reset_index(drop=True)
                workbook_report['模组名称'] = \
                pd.merge(workbook_report, workbook3, on=['库存管理特征'], how='left').reset_index(drop=True)['模组名称200']
                # 过程履历
                add_col = [item for item in workbook_old.columns if item not in workbook_report.columns]
                if not add_col:
                    workbook_old['无过程履历'] = ''
                    add_col = ['无过程履历']
                tgrget_tol = add_col + ['超链码']
                workbook_old1 = workbook_old[tgrget_tol]
                workbook_old1 = workbook_old1.reset_index(drop=True)
                workbook_report = workbook_report.reset_index(drop=True)
                workbook_process = pd.merge(workbook_report, workbook_old1, left_on='超链码', right_on='超链码',
                                            how='left').reset_index(drop=True)
                workbook_process[add_col] = workbook_process[add_col].fillna("空值")
                workbook_process[add_col]=workbook_process[add_col].astype(str)

                if '200最新需求日期' in add_col:
                    add_col.remove('200最新需求日期')
                for col in add_col:
                    workbook_process[col] = ['' if i == '空值' else i for i in workbook_process[col]]
                    workbook_process[col]=workbook_process[col].str.replace('00:00:00','')
                    #workbook_process[col]=workbook_process[col].astype(str).replace('00:00:00','')
                process = workbook_process[add_col].astype(str)
                #process = workbook_process[add_col]#.astype(str)  # 过程履历字段
                # 拉cpjt模组数据
                workbook_module200 = workbook200[~workbook200['库存管理特征'].isin(workbook_item['项目号'])][['项目编号', '库存管理特征']]
                workbook_module200 = workbook_module200[workbook_module200['库存管理特征'].str.contains('无模组号') == False]
                workbook_module200 = workbook_module200.drop_duplicates().reset_index(drop=True)
                workbook_module200 = pd.merge(workbook_module200, workbook_item, left_on='项目编号', right_on='项目号', how='left')[
                    ['项目号', '库存管理特征', '系列']]
                # over_module=list(workbook_module200['库存管理特征'])
                for i in range(len(workbook_module200)):
                    workbook_module200['库存管理特征'][i] = str(workbook_module200['库存管理特征'][i])[::-1].replace('-', '嘬', 1)[::-1]
                # 需求齐料时间
                # 读取被写入的Excel工作簿
                # book = load_workbook('新生成物控表.xlsx')
                app = App(visible=False, add_book=False)
                app.display_alerts = False
                app.screen_updating = False
                wb = app.books.open(book_name + ".xlsx")
                wb1 = app.books.open(entry3.get())
                for i in list(workbook_oldall):
                    if i not in list(workbook_nowall):
                        sheet1 = wb1.sheets[i]
                        sheet2 = wb.sheets['整体数据汇总']
                        #sheet=wb.sheets.add(i)
                        #sheet1.api.Copy(sheet2.api)
                        sheet1.api.Copy(Before=sheet2.api)
                        #sum2sht.api.Copy(Before=sum2sht.api)
                        #sheet1.used_range.copy(wb.sheets[i].range('A1'))
                sht = wb.sheets["物料管控表"]
                '''
                workbook_report['最新需求日期-200'] = workbook_report['最新需求日期-200'].fillna(default_date)
                workbook_report['最新需求日期-200'] = pd.to_datetime(workbook_report['最新需求日期-200'], errors='coerce').dt.strftime(
                    '%Y-%m-%d')
                workbook_report['最新需求日期-200'] = workbook_report[['最新需求日期-200']].astype(str)
                workbook_report['最新需求日期-200'] = ['' if i == '1990-01-01' else i
                                                 for i in workbook_report['最新需求日期-200']]
                #sht.api.Columns(24).Insert()  # 26列新插入一列
                sht.range('BG2').expand('table').value = workbook_report[['最新需求日期-200']].values.tolist()
                '''
                sht.range('D2').expand('table').value = workbook_report[['模组名称']].values.tolist()
                sht.range('D1').color = 'DEA900'
                '''
                sht.range('BG1').color = 'DEA900'
                sht.range('BG1').value = '200最新需求日期'
                sht.range('BG1').column_width = 12
                '''
                # sht.range('BB1').color = 'DEA900'
                t = len(workbook_report['当前状态']) + 1
                '''
                sht.range('BG2:BG%d' % t).api.NumberFormat = "yyyy-mm-dd"
                '''
                # 写入过程履历
                #process = process.fillna('你哈')
                c=49
                for i in list(process.columns):
                    #process[i]=process[i].fillna('你好')
                    sht.api.Columns(50).Insert()  # 37列新插入一列
                    #sht.range(get_column_letter(c+1)+str(1)).value = str(i)
                    c=c+1
                #sht.range('AW1').options(pd.DataFrame, index=False).value = process
                sht.range('AX2').expand('table').value = process.values.tolist()
                colnum = len(process.columns) + 49
                end = get_column_letter(colnum)
                n = str(1)
                sht.range('AX1:' + end + n).color = 'FF0000'
                # sht.range('BD:'+end).columns.autofit()#列宽自适应
                sht.range('AX:' + end).column_width = 15

                c = 49
                for i in list(process.columns):
                    # process[i]=process[i].fillna('你好')
                   # sht.api.Columns(49).Insert()  # 37列新插入一列
                    sht.range(get_column_letter(c + 1) + str(1)).value = str(i)
                    c = c + 1
                # 插入未下单模组
                sht_item = wb.sheets["整体数据汇总"]
                # if len(workbook_module200)>=1:
                #    m=len(workbook_module200)-1
                #   for i in range(m):
                #      print('你好')
                n = len(workbook_item) + 5
                o = n + len(workbook_module200) - 1
                sht_item.range('B%d' % n).expand('table').value = workbook_module200[['库存管理特征']].values.tolist()
                sht_item.range('C%d' % n).expand('table').value = workbook_module200[['系列']].values.tolist()
                sht_item.range('B%d' % n + ':B%d' % o).api.Font.ColorIndex = 3
                sht_item.range('C%d' % n + ':C%d' % o).api.Font.ColorIndex = 3
                # 排序
                # sht_item.range('a5',(rows,columns)).api.Sort(Key1=sht.range('c5').api,Key2=sht.range('B5').api, Order1 = 1,Order2 = 1,Orientation=1)
                sht_item.range('C5:C%d' % o).api.Replace("系列总计", "嘬系列总计")
                sht_item.range('A5:AX%d' % o).api.Sort(Key1=sht_item.range('c5').api, Order1=1, Key2=sht_item.range('b5').api,
                                                       Order2=1, Orientation=1)
                sht_item.range('B5:C%d' % o).api.Replace("嘬", "-")
                app.display_alerts = True
                app.screen_updating = True
                wb.save()
                # 关闭excel程序
                wb.close()
                wb1.close()
                app.quit()
                errorscrem.insert(INSERT, '\n***************执行正常***************')
            if len(workbook_report) <= 0:
                errorscrem.insert(INSERT, '\n最新物控表无数据，不予加工！！！！！！！！！！！！！！！！！！！！！！！！！！！！！！！')
        except Exception as f:
            #print('异常信息为:', e)  # 异常信息为: division by zero
            errorscrem.insert(INSERT, '\n***************程序报错，异常信息为:' + traceback.format_exc())
            print('\n***************程序报错，异常信息为:' + traceback.format_exc())

    btn4['command']=excel
    newWindow.mainloop()
add_deal['command']=createNewWindow
window.mainloop()
